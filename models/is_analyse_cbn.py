# -*- coding: utf-8 -*-
from odoo import models,fields,api
import odoo.tools as tools
from datetime import datetime, timedelta
#import json 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, borders
from openpyxl.styles.borders import Border
from openpyxl.worksheet.page import PageMargins
import base64
import logging
_logger = logging.getLogger(__name__)

#TODO : 
# - Générer un fichie CSV / Excel pour la valorisation (Ajouter un bouton en haut à droite pour télécharger le fichier généré)
# - Ajouter le bouton pour rafraichier une ligne
# - Ajouter un bouton pour supprimer une ligne (juste à l'affichage)
# - Rechercher la liste des fournisseurs uniquement sur la base article pour gagner 1 à 2s par recherche
# - Supprumer les données inutiles transférr au code javascip (ex : supprimer les dictionnaires remplacés par des liste )
# - Continuer à rechercher si il est possible d'envoyer un dictionnaire dans un foreach (t-key = xxx_index)
# - Optimiser le temps de traitement (convertir dict en list) et supprimer les données tranférées inutilement (dict)
# - Faire fonctionner le cache lors de la modificiation des ODs
# - Mesurer le temps de traitement de chaque partie pour optimiser le temps


class product_product(models.Model):
    _inherit = "product.product"

    def get_analyse_cbn(self, 
            ok=False,
            code_pg=False, 
            gest=False, 
            cat=False, 
            moule=False, 
            projet=False, 
            client=False, 
            fournisseur=False, 
            semaines=False, 
            type_cde=False, 
            type_rapport=False, 
            calage=False, 
            valorisation=False,
            product_id=False,
    ):
        mem_product_id = product_id
        cr = self._cr
        debut=datetime.now()
        _logger.info('Début')

        #** set/get var *****************************************************
        debut2=datetime.now()
        if ok:
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_code_pg'     , code_pg)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_gest'        , gest)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_cat'         , cat)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_moule'       , moule)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_projet'      , projet)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_client'      , client)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_fournisseur' , fournisseur)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_semaines'    , semaines)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_type_cde'    , type_cde)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_type_rapport', type_rapport)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_calage'      , calage)
            self.env['is.mem.var'].set(self._uid, 'analyse_cbn_valorisation', valorisation)
        else:
            code_pg      = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_code_pg')
            gest         = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_gest')
            cat          = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_cat')
            moule        = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_moule')
            projet       = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_projet')
            client       = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_client')
            fournisseur  = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_fournisseur')
            semaines     = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_semaines')
            type_cde     = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_type_cde')
            type_rapport = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_type_rapport')
            calage       = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_calage')
            valorisation = self.env['is.mem.var'].get(self._uid, 'analyse_cbn_valorisation')
        semaines = int(semaines)
        _logger.info("set/get var (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())


        semaines = semaines or 18

        #** Valeur par défaut **************************************************
        code_pg       = code_pg  or ''
        gest          = gest           or ''
        cat           = cat            or ''
        moule         = moule          or ''
        projet        = projet         or ''
        client        = client         or ''
        fournisseur   = fournisseur    or ''
        type_cde      = type_cde       or ''
        type_rapport  = type_rapport   or 'Fabrication'
        calage        = calage         or 'Date de fin'
        #***********************************************************************

        # ** Filtre pour les requêtes ******************************************
        debut2=datetime.now()
        filtre=""
        if product_id:
            filtre=filtre+" and pp.id="+str(product_id)+" "

        if code_pg:
            filtre=filtre+" and pt.is_code ilike '"+code_pg+"%' "
        if type_rapport=="Achat":
            filtre=filtre+" and pt.purchase_ok=true "
        else:
            filtre=filtre+" and pt.purchase_ok<>true "
        if cat:
            filtre=filtre+" and ic.name='"+cat+"' "
        if moule:
            moules = moule.split(',')
            res=[]
            for m in moules:
                res.append("'"+str(m)+"'")
            moules=",".join(res)
            moules='('+moules+')'
            filtre=filtre+" and (im.name in "+moules+" or id.name in "+moules+" )"
        if projet:
            filtre=filtre+" and (imp1.name ilike '%"+projet+"%' or imp2.name ilike '%"+projet+"%') "
        if client:
            filtre=filtre+" and rp.is_code='"+client+"' "
        _logger.info("Filtre pour les requêtes (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        # *********************************************************************




        #** Titre, TypeCde, Gest et Fournisseur *******************************
        titre              = self._get_titre(type_rapport,valorisation)
        #cat2id             = self._get_cat2id()                          # Catégories
        if valorisation=="Oui":
            debut2=datetime.now()
            Couts = self._get_Couts()                           # Coûts
            _logger.info("TypeCde (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        TypeCde            = self._get_TypeCde(type_rapport)             # Type de commande d'achat
        _logger.info("TypeCde (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        select_gest        = self._get_select_gest(filtre,fournisseur)   # Liste des gestionnaires
        _logger.info("select_gest (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        select_fournisseur = self._get_select_fournisseur(filtre,gest)   # Liste des fournisseurs
        _logger.info("select_fournisseur (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        # *********************************************************************

        if len(select_fournisseur)==1:
            fournisseur=False

        #Liste de choix *******************************************************
        debut2=datetime.now()
        options = select_gest
        gest_options=[]
        for o in options:
            selected=False
            if o==gest:
                selected=True
            gest_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        options = select_fournisseur
        fournisseur_options=[]
        for o in options:
            selected=False
            if o==fournisseur:
                selected=True
            fournisseur_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        options = [4,8,12,16,18,20,25,30,40,60]
        semaines_options=[]
        for o in options:
            selected=False
            if o==int(semaines):
                selected=True
            semaines_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        options = ['','ferme_uniquement','ferme','ouverte','cadencée']
        type_cde_options=[]
        for o in options:
            selected=False
            if o==type_cde:
                selected=True
            type_cde_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        options = ["Achat", "Fabrication"]
        type_rapport_options=[]
        for o in options:
            selected=False
            if o==type_rapport:
                selected=True
            type_rapport_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        options = ["Date de fin", "Date de début"]
        calage_options=[]
        for o in options:
            selected=False
            if o==calage:
                selected=True
            calage_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        options = ["Non", "Oui"]
        valorisation_options=[]
        for o in options:
            selected=False
            if o==valorisation:
                selected=True
            valorisation_options.append({
                "id": o,
                "name": o,
                "selected": selected,
            })
        _logger.info("Liste de choix (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        # **********************************************************************


        # ** Filtre sur le fournisseur (partner_id) ****************************
        debut2=datetime.now()
        partner_id=False
        if fournisseur:
            tab=fournisseur.split('-')
            SQL="select id from res_partner where is_code='"+tab[0]+"' and is_adr_code='"+tab[1]+"' "
            cr.execute(SQL)
            result = cr.fetchall()
            for row in result:
                partner_id=row[0]
        if gest:
            filtre=filtre+" and ig.name='"+gest+"' "
        if partner_id:
            filtre=filtre+""" 
                and (
                        (   select rp2.id from product_supplierinfo ps inner join res_partner rp2 on ps.partner_id=rp2.id   
                            where ps.product_tmpl_id=pt.id order by ps.sequence limit 1
                        )="""+str(partner_id)+"""
                    )
            """
        filtre=filtre+" AND ic.name not in ('70','71','72','73','74','75','82')"
        _logger.info("Filtre sur le fournisseur (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        # **********************************************************************


        #** Titres des colonnes ***********************************************
        debut2=datetime.now()
        now = datetime.now()
        d = now - timedelta(days=now.weekday())
        TabSemaines={}
        for o in range(0,int(semaines)):
            #week = d.isocalendar().week
            key = d.strftime('%Y%m%d')
            TabSemaines[key] = {
                "key": key,
                "semaine": d.strftime("S%V"),
                "date": d.strftime("%d.%m"),
            }
            d = d + timedelta(days=7)
        #print(json.dumps(TabSemaines, indent = 4))
        _logger.info("Titres des colonnes (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************

        #** Stock *************************************************************
        #TabSemaines = self._get_TabSemaines(semaines)   # Tableau des semaines

        debut2=datetime.now()
        StocksA     = self._get_Stocks('f')              # StocksA (control_quality=f)
        _logger.info("StocksA (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        StocksQ     = self._get_Stocks('t')              # StocksQ (control_quality=t)
        _logger.info("StocksQ (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        StocksSecu  = self._get_StocksSecu()             # StocksSecu
        _logger.info("StocksSecu (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        Fournisseurs, Delai_Fournisseurs = self._get_Fournisseurs_Delai_Fournisseurs() # Fournisseurs + Delai_Fournisseurs
        _logger.info("Fournisseurs (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************


        #** Résultat **********************************************************
        debut2=datetime.now()
        result1 = self._get_FS_SA(filtre)
        result2 = self._get_CF_CP(filtre)
        result3 = self._get_FL(filtre)
        result4 = self._get_FM(filtre)
        result5 = self._get_SF(filtre)
        result = result1+result2+result3+result4+result5
        if valorisation=="Oui":
            result+=self._get_stock(filtre)
            
        res={}
        for row in result:
            product_id = row["product_id"]
            code    = row["code_pg"]
            if type_rapport=='Achat':
                code_fournisseur=Fournisseurs.get(product_id,"0000")
                key="%s/%s"%(code_fournisseur,code)
                k = code_fournisseur+'/'+str(product_id)
                t=TypeCde.get(k)
                if t:
                    Code = "%s / %s / %s (%s)"%(code_fournisseur,code,row["moule"],t)
                else:
                    Code = "%s / %s / %s "%(code_fournisseur,code,row["moule"])
            else:
                key="%s/%s"%(row["moule"],code)
                Code = "%s / %s"%(row["moule"],code)
            cout=0
            if valorisation=="Oui":
                cout = Couts.get(product_id,0)
                Code = "%s (%s)"%(Code, cout)

            if key not in res:
                product_id = row["product_id"]
                product = self.env['product.product'].browse(product_id)
                product_tmpl = product.product_tmpl_id
                if type_rapport=='Achat':
                    Delai=Delai_Fournisseurs.get(product_id,0)
                else:
                    Delai=row["produce_delay"]
                res[key] = {
                    "key"        : key,
                    "Code"       : Code,
                    "product_id" : product_id,
                    "product_tmpl_id": product_tmpl.id,
                    "code_pg"    : row["code_pg"],
                    "designation": row["designation"],
                    "cout"       : cout or 0,
                    "lot_mini"   : row["lot_mini"],
                    "multiple"   : row["multiple"],
                    "StockSecu"  : row["is_stock_secu"],
                    "Delai"      : Delai,
                    "StockA"     : int(StocksA.get(product_id,0) or 0),
                    "StockQ"     : int(StocksQ.get(product_id,0) or 0),
                    "typeod"     : {},
                }

                #Ajout des colonnes pour le stock *****************************
                res[key]["typeod"]["90-Stock"]={
                    "key"        : "90-Stock",
                    "typeod"     : "90-Stock",
                    "name_typeod": "Stock",
                    "cols"  : {}
                }
                for d in TabSemaines:
                    res[key]["typeod"]["90-Stock"]["cols"][d]={
                        "key"     : d,
                        "qt_signe": 0,
                        "qt_txt"  : "",
                        "od"       : []
                    }
                #**************************************************************

                #Ajout des colonnes pour le stock valorisé ********************
                if valorisation=="Oui":
                    lig="92-Stock Valorisé"
                    res[key]["typeod"][lig]={
                        "key"        : lig,
                        "typeod"     : lig,
                        "name_typeod": "Valorisation",
                        "cols"  : {}
                    }
                    for d in TabSemaines:
                        res[key]["typeod"][lig]["cols"][d]={
                            "key"     : d,
                            "qt": 0,
                            "qt_signe": 0,
                            "qt_txt"  : "",
                            "od"       : []
                        }
                #**************************************************************

            typeod = row["typeod"]
            key2   = self._get_name_typeod(typeod)
            if key2 not in res[key]["typeod"]:
                res[key]["typeod"][key2] = {
                    "key"        : key2,
                    "typeod"     : row["typeod"],
                    "name_typeod": key2[3:],
                    "cols"  : {}
                }
                for d in TabSemaines:
                    res[key]["typeod"][key2]["cols"][d]={
                        "key"   : d,
                        "qt"    : 0,
                        "qttxt" : "",
                        "od"    : {},
                        #"ids"   : [],
                    }
            if calage=='' or calage=='Date de fin':
                DateLundi=self.datelundi(row["date_fin"], TabSemaines)
            else:
                DateLundi=self.datelundi(row["date_debut"], TabSemaines)
            if DateLundi:
                if DateLundi in TabSemaines:
                    v=round(row["qt"],6)
                    if row["typeod"]=='FL' and v<0:
                        v=0
                    qt = res[key]["typeod"][key2]["cols"][DateLundi]["qt"]+v
                    #qt = res[key]["typeod"][key2]["cols"][DateLundi]["qt"]+round(row["qt"],6)
                    color_qt = self._get_color_qt(key2,qt)
                    qt_signe = qt * self._get_sens(typeod)
                    qt_txt=""
                    if qt>0:
                        qt_txt = round(qt_signe)

                    #if row["numod"] not in res[key]["typeod"][key2]["cols"][DateLundi]["ids"]:
                    #    res[key]["typeod"][key2]["cols"][DateLundi]["ids"].append(row["numod"])

                    #** Afficher l'icon trash si un seul OD et du type accecpté *********
                    trash=False
                    if qt_txt!="" and key2[3:] in ['FS','SA']:
                        trash=True
                    #********************************************************************
                    numod = row["numod"]
                    v={
                        "numod": numod,
                        "name" : row["name"],
                        "qt"   : round(row["qt"],4),
                        "trash": trash,
                    }
                    res[key]["typeod"][key2]["cols"][DateLundi]["od"][numod] = v


                    #if row["name"] not in res[key]["typeod"][key2]["cols"][DateLundi]["od"]:
                    #    res[key]["typeod"][key2]["cols"][DateLundi]["od"].append(row["name"])
                    #od_txt = ", ".join(res[key]["typeod"][key2]["cols"][DateLundi]["od"])

                    #** Afficher l'icon trash si un seul OD et du type accecpté *********
                    # trash=False
                    # if qt_txt!="":
                    #     if len(res[key]["typeod"][key2]["cols"][DateLundi]["od"])==1:
                    #         if key2[3:] in ['FL','FS','SA']:
                    #             trash=True
                    # #********************************************************************

                    res[key]["typeod"][key2]["cols"][DateLundi].update({
                        "qt"      : qt,
                        "color_qt": color_qt,
                        "qt_txt"  : qt_txt,
                        "qt_signe": qt_signe,
                        #"od_txt"  : od_txt,
                        #"trash"   : trash,
                    })
            #res[key]["typeod"][key2]["colslist"] = list(res[key]["typeod"][key2]["cols"].values())
        _logger.info("Résultat (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************

        #Convertir un dictionnaire en list (pour owl) avec un tri sur les clés => Voir pour optimiser le temps de traitement
        debut2=datetime.now()
        for key in res:
            res[key]["typeodlist"]=[]
            for line in dict(sorted(res[key]["typeod"].items())):
                res[key]["typeodlist"].append(res[key]["typeod"][line])
            res[key]["rowspan"]    = len(res[key]["typeodlist"])
        _logger.info("Convertir un dictionnaire en list (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************

        #** Calcul du total des besoins par date ******************************
        debut2=datetime.now()
        totaux={}
        for p in res:
            test=False
            totaux[p]={}
            for t in res[p]["typeod"]:
                if t not in["90-Stock","92-Stock Valorisé"]:
                    typeod = res[p]["typeod"][t]["typeod"]
                    for c in TabSemaines:
                        qt=res[p]["typeod"][t]["cols"][c]["qt"]
                        qt_signe = qt * self._get_sens(typeod)
                        if qt!=0:
                            test=True
                        if c not in totaux[p]:
                            totaux[p][c]=0
                        totaux[p][c]+=qt_signe

                if valorisation=="Oui":
                    if t=="90-Stock":
                        if res[p]['StockA'] or res[p]['StockQ']:
                            test=True
            res[p]['test']=test #Permet de détecter si l'article a du stok ou des besoins
        _logger.info("Calcul du total des besoins par date (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************

        #** Calcul du stock cumulé par date ***********************************
        debut2=datetime.now()
        for p in totaux:
            ct=0
            cumul=0
            for c in TabSemaines:
                if ct==0:
                    if valorisation=="Oui":
                        cumul = res[p]["StockA"]+res[p]["StockQ"]
                    else:
                        cumul = res[p]["StockA"]-res[p]["StockSecu"]
                cumul+=totaux[p][c]
                val = cumul*res[p]["cout"]
                if val<0:
                    val=0

                color="Gray"
                if cumul<0:
                    color='Red'
                res[p]["typeod"]["90-Stock"]["cols"][c]["qt_signe"]=cumul
                res[p]["typeod"]["90-Stock"]["cols"][c]["color_qt"]=color
                res[p]["typeod"]["90-Stock"]["cols"][c]["qt_txt"] = round(res[p]["typeod"]["90-Stock"]["cols"][c]["qt_signe"])
                if valorisation=="Oui":
                    res[p]["typeod"]["92-Stock Valorisé"]["cols"][c]["qt_txt"] = round(val)
                ct+=1
        _logger.info("Calcul du stock cumulé par date (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************


        #** Supprime les articles sans stock et sans besoins ******************
        res_epure = {}
        for p in res:
            if res[p]['test']:
                res_epure[p] = res[p]
        #res_epure = res
        #**********************************************************************


        #** Ajout de la couleur des lignes ************************************
        debut2=datetime.now()
        sorted_dict = dict(sorted(res_epure.items())) 
        trcolor=""
        for k in sorted_dict:
            if trcolor=="#ffffff":
                trcolor="#f2f3f4"
            else:
                trcolor="#ffffff"
            if mem_product_id:
                trcolor="#00FAA2"
            trstyle="background-color:%s"%(trcolor)
            sorted_dict[k]["trstyle"] = trstyle
        #lines = list(sorted_dict.values())
        _logger.info("Ajout de la couleur des lignes (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************


        #** Création fichier Excel ********************************************
        excel_attachment_id = False
        if valorisation=="Oui":
            #** Création du workbook ******************************************
            user  = self.env['res.users'].browse(self._uid)
            name  = 'analyse-cbn-%s.xlsx'%user.login
            path = '/tmp/%s'%name
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Analyse CBN"
            #******************************************************************

            #** Ligne d'entête ************************************************
            clr_background  = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type="solid")
            clr_background1 = PatternFill(start_color='98FB98', end_color='98FB98', fill_type="solid")
            clr_background2 = PatternFill(start_color='a0c8f0', end_color='a0c8f0', fill_type="solid")




            cell = ws.cell(row=1, column=1, value="Code PG")
            cell.fill = clr_background
            cell.font = Font(name='Calibri', bold=True)
            cell = ws.cell(row=1, column=2, value="Désignation")
            cell.font = Font(name='Calibri', bold=True)
            cell.fill = clr_background
            for key in sorted_dict:
                line = sorted_dict[key]["typeod"]["92-Stock Valorisé"]["cols"]
                column = 3
                for col in line:
                    def set_cell(ws,column,val,background):
                        cell = ws.cell(row=1, column=column, value=val)
                        #cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = background
                        cell.font = Font(name='Calibri', bold=True)
 
                    date = sorted_dict[key]["typeod"]["90-Stock"]["cols"][col]["key"]
                    val =  "Stock\n%s"%(datetime.strptime(date, '%Y%m%d').strftime('%d/%m/%Y'))
                    set_cell(ws,column,val,clr_background1)
                    column+=1

                    date = sorted_dict[key]["typeod"]["92-Stock Valorisé"]["cols"][col]["key"]
                    val =  "Val\n%s"%(datetime.strptime(date, '%Y%m%d').strftime('%d/%m/%Y'))
                    set_cell(ws,column,val,clr_background2)
                    column+=1
                break
            #******************************************************************

            #** Contenu *******************************************************
            row = 2
            for key in sorted_dict:
                code        = sorted_dict[key]["code_pg"]
                designation = sorted_dict[key]["designation"]
                d = ws.cell(row=row, column=1, value=code)
                d = ws.cell(row=row, column=2, value=designation)
                line = sorted_dict[key]["typeod"]["92-Stock Valorisé"]["cols"]
                column = 3
                for col in line:
                    qt = sorted_dict[key]["typeod"]["90-Stock"]["cols"][col]["qt_txt"]
                    cell = ws.cell(row=row, column=column, value=qt)
                    cell.number_format = '# ##0'  # Number formatting
                    column+=1

                    qt = sorted_dict[key]["typeod"]["92-Stock Valorisé"]["cols"][col]["qt_txt"]
                    cell = ws.cell(row=row, column=column, value=qt)
                    cell.number_format = '# ##0'  # Number formatting
                    column+=1
                row+=1
            #******************************************************************

            #** Bordures des cellules *****************************************
            border1 = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
            thin_border = Border(left = border1, right = border1, bottom = border1, top = border1)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            #******************************************************************

            #** Ajustement de la largeur des colonnes *************************
            for idx, col in enumerate(ws.columns, 1):
                ws.column_dimensions[get_column_letter(idx)].auto_size = True
            ws.column_dimensions['B'].width=30
            #******************************************************************
  
            #** Mise en page **************************************************
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.oddFooter.center.text = "Page &[Page] of &N"
            ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.7, header=0, footer=0.3)
            ws.print_title_rows = '1:1' # the first two rows 
            #******************************************************************

            #** Fixer les lignes et les colonnes ******************************
            cel = ws['C2']
            ws.freeze_panes = cel
            #******************************************************************

            #** Enregistrement du workbook ************************************
            workbook.save(path)
            #******************************************************************

            # ** Creation ou modification de la pièce jointe ******************
            attachment_obj = self.env['ir.attachment']
            attachments = attachment_obj.search([('res_id','=',user.id),('name','=',name)])
            xlsx = open(path,'rb').read()
            vals = {
                'name':        name,
                'type':        'binary',
                'res_id':      user.id,
                'datas':       base64.b64encode(xlsx),
            }
            attachment_id=False
            if attachments:
                for attachment in attachments:
                    attachment.write(vals)
                    attachment_id=attachment.id
            else:
                attachment = attachment_obj.create(vals)
                attachment_id=attachment.id
            excel_attachment_id = attachment_id
            #*******************************************************************
        #**********************************************************************




        #** Sauvegarde du résultat ********************************************
        #TODO : Cela ne semble pas pertinent, car Odoo met moins de 2s à générer le résultat, 
        # mais le navigateur 6s à traiter les 8Mo d'informations
        # x = json.dumps(sorted_dict)
        # print("Taille du json : %.1fMo"%(len(x)/1024/1024))
        # self.env['is.mem.var'].set(self._uid, 'analyse_cbn_dict', x)
        # _logger.info("enregistrement json (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())
        #**********************************************************************

        duree = datetime.now()-debut
        _logger.info("Fin (durée=%.2fs)"%(datetime.now()-debut).total_seconds())
        res={
            "titre"       : titre,
            "dict"        : sorted_dict,
            "date_cols"   : list(TabSemaines.values()),
            "code_pg"     : code_pg,
            "gest"        : gest,
            "cat"         : cat,
            "moule"       : moule,
            "projet"      : projet,
            "client"      : client,
            "fournisseur" : fournisseur,
            "semaines"    : semaines,
            "type_cde"    : type_cde,
            "type_rapport": type_rapport,
            "calage"      : calage,
            "valorisation": valorisation,
            "gest_options"        : gest_options,
            "fournisseur_options" : fournisseur_options,
            "semaines_options"    : semaines_options,
            "type_cde_options"    : type_cde_options,
            "type_rapport_options": type_rapport_options,
            "calage_options"      : calage_options,
            "valorisation_options": valorisation_options,
            "excel_attachment_id" : excel_attachment_id,
        }
        return res


    def _get_titre(self,type_rapport,valorisation):
        titre="Suggestions de fabrication"
        if type_rapport=="Achat":
            titre="Suggestions d'achat"
        if valorisation=="Oui":
            titre="Valorisation stock fabrication"
            if type_rapport=="Achat":
                titre="Valorisation stock achat"
        return titre


    def _get_cat2id(self):
        cr = self._cr
        SQL="""
            select id, name
            from is_category 
            where id>0
        """
        cr.execute(SQL)
        result = cr.fetchall()
        cat2id={}
        for row in result:
            cat2id[row[1]]=row[0]
        return cat2id


    def _get_Couts(self):
        cr = self._cr
        SQL="""
            select name, cout_act_total
            from is_cout 
        """
        cr.execute(SQL)
        result = cr.fetchall()
        Couts={}
        for row in result:
            Couts[row[0]]=row[1]
        return Couts


    def _get_TypeCde(self,type_rapport):
        cr = self._cr
        TypeCde={}
        if type_rapport=="Achat":
            SQL="""
                SELECT 
                    cof.type_commande, 
                    cof.partner_id, 
                    rp.is_code, 
                    rp.is_adr_code, 
                    cofp.product_id
                FROM is_cde_ouverte_fournisseur cof inner join is_cde_ouverte_fournisseur_product cofp on cofp.order_id=cof.id
                                                    inner join res_partner rp on cof.partner_id=rp.id
            """
            cr.execute(SQL)
            result = cr.fetchall()
            for row in result:
                cle=('0000'+row[2])[-4:]+'-'+row[3]+'/'+str(row[4])
                TypeCde[cle]=row[0]
            SQL="""
                SELECT 
                    rp.is_code, 
                    rp.is_adr_code, 
                    product_id
                FROM is_cde_ferme_cadencee cfc inner join res_partner rp on cfc.partner_id=rp.id
            """
            cr.execute(SQL)
            result = cr.fetchall()
            for row in result:
                cle=('0000'+row[0])[-4:]+'-'+row[1]+'/'+str(row[2])
                TypeCde[cle]=u'cadencée'
        return TypeCde


    def _get_select_gest(self,filtre,fournisseur):
        cr = self._cr
        SQL="""
            SELECT distinct ig.name gest
            FROM product_product pp inner join product_template      pt   on pp.product_tmpl_id=pt.id

                                    left outer join is_mold          im   on pt.is_mold_id=im.id
                                    left outer join is_dossierf      id   on pt.is_dossierf_id=id.id
                                    inner join      is_gestionnaire  ig   on pt.is_gestionnaire_id=ig.id
                                    left outer join is_category      ic   on pt.is_category_id=ic.id
                                    left outer join is_mold_project  imp1 on im.project=imp1.id
                                    left outer join is_mold_project  imp2 on id.project=imp2.id
                                    left outer join res_partner      rp   on pt.is_client_id=rp.id
            WHERE pp.id>0 """+filtre+"""
        """
        if fournisseur:
            code=fournisseur.split('-')[0]
            SQL=SQL+""" 
                and (
                    (
                        select rp2.is_code 
                        from product_supplierinfo ps inner join res_partner rp2 on ps.partner_id=rp2.id   
                        where ps.product_tmpl_id=pt.id order by ps.sequence,ps.id limit 1
                    )='"""+code+"""'
                )

            """
        SQL=SQL+""" ORDER BY ig.name """
        cr.execute(SQL)
        result = cr.fetchall()
        SelectGestionnaires={}
        for row in result:
            SelectGestionnaires[row[0]]=1
        select_gest=[]
        select_gest.append('')
        for x in SelectGestionnaires:
            select_gest.append(x)
        return select_gest


    def _get_select_fournisseur(self,filtre,gest=False):
        cr = self._cr
        if gest:
            filtre=filtre+" and ig.name='"+gest+"' "

        #** Recherche des articles ayant des données ***************************
        _logger.info("filtre=%s"%(filtre))
        debut2=datetime.now()
        r1=self._get_FS_SA(filtre)
        _logger.info("_get_FS_SA (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        r2=self._get_CF_CP(filtre)
        _logger.info("_get_CF_CP (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        r3=self._get_FL(filtre)
        _logger.info("_get_FL (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        r4=self._get_FM(filtre)
        _logger.info("_get_FM (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        debut2=datetime.now()
        r5=self._get_SF(filtre)
        _logger.info("_get_SF (durée=%.2fs)"%(datetime.now()-debut2).total_seconds())

        product_ids=[]
        for row in (r1+r2+r3+r4+r5):
            product_id=str(row["product_id"])
            if product_id not in product_ids:
                product_ids.append(product_id)
        #***********************************************************************

        SQL="""
            SELECT 
                (   select concat(rp2.is_code,'-',rp2.is_adr_code)
                    from product_supplierinfo ps inner join res_partner rp2 on ps.partner_id=rp2.id   
                    where ps.product_tmpl_id=pt.id order by ps.sequence,ps.id limit 1
                ) code_fournisseur

            FROM product_product pp inner join product_template      pt   on pp.product_tmpl_id=pt.id
                                    left outer join is_mold          im   on pt.is_mold_id=im.id
                                    left outer join is_dossierf      id   on pt.is_dossierf_id=id.id
                                    inner join      is_gestionnaire  ig   on pt.is_gestionnaire_id=ig.id
                                    left outer join is_category      ic   on pt.is_category_id=ic.id
                                    left outer join is_mold_project  imp1 on im.project=imp1.id
                                    left outer join is_mold_project  imp2 on id.project=imp2.id
                                    left outer join res_partner      rp   on pt.is_client_id=rp.id
          WHERE pp.id>0 """+filtre+"""
        """
        if len(product_ids):
            SQL=SQL+" AND pp.id in ("+','.join(product_ids)+") "
        cr.execute(SQL)
        result = cr.fetchall()
        select_fournisseur=[]
        select_fournisseur.append('')
        for row in result:
            cle=row[0]
            if cle and cle not in select_fournisseur:
                select_fournisseur.append(cle)
        select_fournisseur.sort()
        return select_fournisseur


    def _get_FS_SA(self,filtre):
        cr = self._cr
        SQL="""
            SELECT 
                mp.id as numod, 
                mp.start_date as date_debut, 
                mp.start_date_cq as date_fin, 
                mp.quantity as qt, 
                mp.type as typeod, 
                mp.product_id,
                pt.is_code as code_pg, 
                pt.name->>'fr_FR' designation,
                pt.is_stock_secu, 
                pt.produce_delay, 
                pt.lot_mini, 
                pt.multiple,
                pt.is_mold_dossierf as moule,
                mp.name as name,
                ig.name as gest
            FROM mrp_prevision mp inner join product_product      pp   on mp.product_id=pp.id 
                                  inner join product_template     pt   on pp.product_tmpl_id=pt.id
                                  left outer join is_mold         im   on pt.is_mold_id=im.id
                                  left outer join is_dossierf     id   on pt.is_dossierf_id=id.id
                                  left outer join is_gestionnaire ig   on pt.is_gestionnaire_id=ig.id
                                  left outer join is_category     ic   on pt.is_category_id=ic.id
                                  left outer join is_mold_project imp1 on im.project=imp1.id
                                  left outer join is_mold_project imp2 on id.project=imp2.id
                                  left outer join res_partner     rp   on pt.is_client_id=rp.id
            WHERE mp.id>0 """+filtre+""" 
            ORDER BY mp.name
        """
        cr.execute(SQL)
        #result = cr.fetchall()
        result = cr.dictfetchall()
        return result


    def _get_CF_CP(self,filtre):
        cr = self._cr
        SQL="""
            SELECT 
                so.id as numod, 
                sol.is_date_expedition as date_debut, 
                sol.is_date_expedition as date_fin, 
                sol.is_type_commande as typeod, 
                sol.product_id, 
                pt.is_code as code_pg, 
                pt.name->>'fr_FR' designation,
                pt.is_stock_secu, 
                pt.produce_delay, 
                pt.lot_mini, 
                pt.multiple,
                pt.is_mold_dossierf as moule,
                so.name as name,

                sum(sol.product_uom_qty) as qt


            FROM sale_order_line sol inner join sale_order           so   on sol.order_id=so.id 
                               inner join product_product      pp   on sol.product_id=pp.id
                               inner join product_template     pt   on pp.product_tmpl_id=pt.id
                               left outer join is_mold         im   on pt.is_mold_id=im.id
                               left outer join is_dossierf     id   on pt.is_dossierf_id=id.id
                               left outer join is_gestionnaire ig   on pt.is_gestionnaire_id=ig.id
                               left outer join is_category     ic   on pt.is_category_id=ic.id
                               left outer join is_mold_project imp1 on im.project=imp1.id
                               left outer join is_mold_project imp2 on id.project=imp2.id
                               left outer join res_partner     rp   on pt.is_client_id=rp.id
            WHERE sol.id>0 """+filtre+""" and so.state in ('draft','sent')
            GROUP BY 
                so.id, sol.is_date_expedition, sol.is_type_commande, sol.product_id, pt.is_code, pt.name, pt.is_stock_secu, pt.produce_delay,
                pt.lot_mini, pt.multiple, pt.is_mold_dossierf, so.name
            ORDER BY so.name
        """
        cr.execute(SQL)
        result = cr.dictfetchall()
        return result


    def _get_FL(self,filtre):
        cr = self._cr
        # SQL="""
        #     SELECT 
        #         mp.id as numod, 
        #         mp.date_planned_start as date_debut, 
        #         mp.date_planned_start as date_fin, 
        #         sm.product_qty as qt, 
        #         'FL' as typeod, 
        #         sm.product_id, 
        #         pt.is_code as code_pg, 
        #         pt.name->>'fr_FR' designation,
        #         pt.is_stock_secu, 
        #         pt.produce_delay, 
        #         pt.lot_mini, 
        #         pt.multiple,
        #         pt.is_mold_dossierf as moule,
        #         mp.name as name
        #     FROM stock_move sm    inner join product_product       pp   on sm.product_id=pp.id
        #                           inner join product_template      pt   on pp.product_tmpl_id=pt.id
        #                            inner join mrp_production       mp   on mp.id=sm.production_id
        #                             left outer join is_mold        im   on pt.is_mold_id=im.id
        #                            left outer join is_dossierf     id   on pt.is_dossierf_id=id.id
        #                            left outer join is_gestionnaire ig   on pt.is_gestionnaire_id=ig.id
        #                            left outer join is_category     ic   on pt.is_category_id=ic.id and ic.name!='74'
        #                            left outer join is_mold_project imp1 on im.project=imp1.id
        #                            left outer join is_mold_project imp2 on id.project=imp2.id
        #                            left outer join res_partner     rp   on pt.is_client_id=rp.id
        #     WHERE sm.id>0 """+filtre+""" and production_id>0 and sm.state<>'done' and sm.state<>'cancel'
        #     ORDER BY sm.name 
        # """

        SQL="""
            SELECT 
                mp.id as numod, 
                mp.date_planned_start as date_debut, 
                mp.date_planned_start as date_fin, 
                mp.is_qt_reste_uom as qt, 
                'FL' as typeod, 
                mp.product_id, 
                pt.is_code as code_pg, 
                pt.name->>'fr_FR' designation,
                pt.is_stock_secu, 
                pt.produce_delay, 
                pt.lot_mini, 
                pt.multiple,
                pt.is_mold_dossierf as moule,
                mp.name as name
            FROM mrp_production mp inner join product_product      pp   on mp.product_id=pp.id
                                   inner join product_template      pt   on pp.product_tmpl_id=pt.id
                                   left outer join is_mold        im   on pt.is_mold_id=im.id
                                   left outer join is_dossierf     id   on pt.is_dossierf_id=id.id
                                   left outer join is_gestionnaire ig   on pt.is_gestionnaire_id=ig.id
                                   left outer join is_category     ic   on pt.is_category_id=ic.id 
                                   left outer join is_mold_project imp1 on im.project=imp1.id
                                   left outer join is_mold_project imp2 on id.project=imp2.id
                                   left outer join res_partner     rp   on pt.is_client_id=rp.id
            WHERE mp.state not in ('done','cancel') """+filtre+""" 
            ORDER BY mp.name 
        """
# and ic.name!='74'


        cr.execute(SQL)
        result = cr.dictfetchall()
        return result


    def _get_FM(self,filtre):
        cr = self._cr
        SQL="""
            SELECT 
                bom.id as numod, 
                mp.date_planned_start as date_debut, 
                mp.date_planned_start as date_fin, 
                (bom.product_qty*mp.is_qt_reste_uom) as qt, 
                'FM' as typeod, 
                bom.product_id, 
                pt.is_code as code_pg, 
                pt.name->>'fr_FR' designation,
                pt.is_stock_secu, 
                pt.produce_delay, 
                pt.lot_mini, 
                pt.multiple,
                pt.is_mold_dossierf as moule,
                mp.name as name,
                mp.is_qt_reste_uom,
                bom.product_qty
            FROM is_mrp_production_bom bom   inner join mrp_production mp on bom.production_id=mp.id
                            inner join product_product      pp   on bom.product_id=pp.id
                            inner join product_template     pt   on pp.product_tmpl_id=pt.id
                            left outer join is_mold         im   on pt.is_mold_id=im.id
                            left outer join is_dossierf     id   on pt.is_dossierf_id=id.id
                            left outer join is_gestionnaire ig   on pt.is_gestionnaire_id=ig.id
                            left outer join is_category     ic   on pt.is_category_id=ic.id
                            left outer join is_mold_project imp1 on im.project=imp1.id
                            left outer join is_mold_project imp2 on id.project=imp2.id
                            left outer join res_partner     rp   on pt.is_client_id=rp.id

            WHERE mp.state='draft' """+filtre+""" and bom.is_cbn='t'
            ORDER BY mp.name
        """
        cr.execute(SQL)
        result = cr.dictfetchall()
        return result


    def _get_SF(self,filtre):
        cr = self._cr
        SQL="""
            SELECT 
                po.id as numod, 
                
                -- pol.date_planned as date_debut, 
                -- pol.date_planned as date_fin, 

                pol.date_planned + interval '2 hours' as date_debut, 
                pol.date_planned + interval '2 hours' as date_fin, 

                sm.product_qty as qt, 
                'SF' as typeod, 
                pol.product_id, 
                pt.is_code as code_pg, 
                pt.name->>'fr_FR' designation,
                pt.is_stock_secu, 
                pt.produce_delay, 
                pt.lot_mini, 
                pt.multiple,
                pt.is_mold_dossierf as moule,
                po.name as name
            FROM purchase_order_line pol left outer join purchase_order po    on pol.order_id=po.id 
                                     inner join product_product           pp    on pol.product_id=pp.id
                                     inner join product_template          pt    on pp.product_tmpl_id=pt.id
                                     inner join stock_move                sm    on pol.id=sm.purchase_line_id
                                     left outer join is_mold              im    on pt.is_mold_id=im.id
                                     left outer join is_dossierf          id    on pt.is_dossierf_id=id.id
                                     left outer join is_gestionnaire      ig    on pt.is_gestionnaire_id=ig.id
                                     left outer join is_category          ic    on pt.is_category_id=ic.id 
                                     left outer join is_mold_project      imp1  on im.project=imp1.id
                                     left outer join is_mold_project      imp2  on id.project=imp2.id
                                     left outer join res_partner          rp    on pt.is_client_id=rp.id
              WHERE pol.id>0 """+filtre+""" and pol.state<>'draft' and pol.state<>'done' and pol.state<>'cancel'
                    and sm.state in ('draft','waiting','confirmed','assigned')
              ORDER BY pol.name
        """
        # and ic.name not in ('70','72','73','74')
        cr.execute(SQL)
        result = cr.dictfetchall()
        return result


    def _get_TabSemaines(self,nb_semaines):
        date=datetime.now()
        jour=date.weekday()
        date = date - timedelta(days=jour)
        TabSemaines=[]
        for i in range(0,int(nb_semaines)):
            d=date.strftime('%Y%m%d')
            TabSemaines.append(d)
            date = date + timedelta(days=7)
        return TabSemaines


    def _get_Stocks(self,control_quality='f'):
        cr = self._cr
        SQL="""
            select sq.product_id, sum(sq.quantity) as qt
            from stock_quant sq inner join stock_location sl on sq.location_id=sl.id 
            where 
                sl.usage='internal' and 
                sl.active='t' and 
                sl.control_quality='"""+control_quality+"""' 
            group by sq.product_id 
            order by sq.product_id
        """
        cr.execute(SQL)
        result = cr.fetchall()
        Stocks={}
        for row in result:
            Stocks[row[0]]=row[1]        
        return Stocks


    def _get_StocksSecu(self):
        cr = self._cr
        SQL="""
            select pp.id, pt.is_stock_secu as qt
            from product_product pp inner join product_template pt on pp.product_tmpl_id=pt.id 
        """
        cr.execute(SQL)
        result = cr.fetchall()
        StocksSecu={}
        for row in result:
            StocksSecu[row[0]]=row[1]
        return StocksSecu


    def _get_Fournisseurs_Delai_Fournisseurs(self):
        cr = self._cr
        SQL="""
            select 
                pp.id      as product_id, 
                rp.is_code as is_code, 
                rp.name    as name, 
                ps.delay   as delay, 
                rp.is_adr_code
            from product_product pp, product_supplierinfo ps, res_partner rp 
            where pp.product_tmpl_id=ps.product_tmpl_id and ps.partner_id=rp.id 
            order by ps.sequence, ps.id
        """
        cr.execute(SQL)
        result = cr.fetchall()
        Fournisseurs={}
        Delai_Fournisseurs={}
        for row in result:
            name=('0000'+row[1])[-4:]+'-'+row[4]
            if name=='':
                name=row[2]
            Fournisseurs[row[0]]=name
            Delai_Fournisseurs[row[0]]=row[3]
        return [Fournisseurs,Delai_Fournisseurs]


    def _get_stock(self,filtre):
        cr = self._cr
        SQL="""
            SELECT 
                pt.id as numod, 
                '' as date_debut, 
                '' as date_fin, 
                0 as qt, 
                'stock' as typeod, 
                pp.id as product_id, 
                pt.is_code as code_pg, 
                pt.name->>'fr_FR' designation,
                pt.is_stock_secu, 
                pt.produce_delay, 
                pt.lot_mini, 
                pt.multiple,
                pt.is_mold_dossierf as moule,
                '' as name
            FROM product_product pp inner join product_template         pt    on pp.product_tmpl_id=pt.id
                                   left outer join is_mold              im    on pt.is_mold_id=im.id
                                   left outer join is_dossierf          id    on pt.is_dossierf_id=id.id
                                   left outer join is_gestionnaire      ig    on pt.is_gestionnaire_id=ig.id
                                   left outer join is_category          ic    on pt.is_category_id=ic.id 
                                   left outer join is_mold_project      imp1  on im.project=imp1.id
                                   left outer join is_mold_project      imp2  on id.project=imp2.id
                                   left outer join res_partner          rp    on pt.is_client_id=rp.id
            WHERE pt.id>0 """+filtre+"""
            ORDER BY pt.is_code
        """
        # and ic.name not in ('70','72','73','74')
        cr.execute(SQL)
        result = cr.dictfetchall()
        return result


    def _get_name_typeod(self,typeod):
        "Les chiffres permettent de trier les lignes"
        t={
            "ferme"        : "10-CF",
            "previsionnel" : "20-CP",
            "FL"           : "30-FL",
            "FM"           : "40-FM",
            "SF"           : "50-SF",
            "ft"           : "60-FT",
            "fs"           : "70-FS",
            "sa"           : "80-SA",
            "stock"        : "99-Stock",
        }
        name=""
        if typeod in t:
            name = t[typeod]
        return name


    def _get_sens(self,typeod):
        "Permet de déterminer le sens dans le calcul du stock"
        t={
            "ferme"        : -1,
            "previsionnel" : -1,
            "FL"           : 1,
            "FM"           : -1,
            "SF"           : 1,
            "fs"           : 1,
            "ft"           : -1,
            "sa"           : 1,
            "stock"        : 1,
        }
        sens=1
        if typeod in t:
            sens = t[typeod]
        return sens


    def datelundi(self,date,TabSemaines): 
        firsdate = list(TabSemaines.keys())[0]
        if date=='':
            return firsdate
        if date:
            jour=date.weekday()
            date = date - timedelta(days=jour)
            date = date.strftime('%Y%m%d')
            if date<firsdate:
                date=firsdate
        return date or ''


    def _get_color_qt(self,name_typeod,val=0): 
        name_typeod=name_typeod[3:]
        t={
            "CF"    : "DarkRed",
            "CP"    : "DarkGreen",
            "FL"    : "DarkMagenta",
            "FM"    : "#000000",
            "SF"    : "DarkMagenta",
            "FS"    : "DarkBlue",
            "FT"    : "#000000",
            "SA"    : "DarkBlue",
            "Stock" : "Black",
        }
        color="Gray"
        if name_typeod in t:
            color=t[name_typeod]
        if name_typeod=='Stock' and val<0:
            color='Red'
        return color


        #             #** Calcul du stock valorisé ***************************
        #             if Tab[key]['TypeOD']=='92-Stock Valorisé':
        #                 if i==0:
        #                     Stock=Tab[key][0]+Qt;
        #                 else:
        #                     Stock=TabSho[1+i][lig]+Qt;
        #                 TabSho[2+i][lig]=round(Stock,2)
        #             #*******************************************************
 

    #             # #** Valorisation stock *****************************************
    #             # if valorisation:
    #             #     Cout=0
    #             #     if product_id in Couts:
    #             #         Cout=Couts[product_id]
    #             #     Cle1=code_pg+u'90-Stock'
    #             #     Cle2=code_pg+u'92-Stock Valorisé'
    #             #     if Cle2 not in Tab:
    #             #         Tab[Cle2]={}
    #             #     Tab[Cle2]['Code'] = Code
    #             #     Tab[Cle2]['TypeOD']=u'92-Stock Valorisé'
    #             #     Tab[Cle2][0]=Tab[Cle1][0]*Cout
    #             #     Tab[Cle2][DateLundi]=Tab[Cle1][DateLundi]*Cout
    #             # #***************************************************************
