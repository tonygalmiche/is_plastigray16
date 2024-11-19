# -*- coding: utf-8 -*-
import csv
from odoo import models,fields,api
from odoo.exceptions import ValidationError
from lxml import etree
import xml.etree.ElementTree as ET
import base64
import time
import math
from datetime import date,datetime,timedelta
import openpyxl


_JOURS_SEMAINE=[
    ("1", 'Lundi'),
    ("2", 'Mardi'),
    ("3", 'Mercredi'),
    ("4", 'Jeudi'),
    ("5", 'Vendredi'),
    ("6", 'Samedi'),
    ("7", 'Dimanche'),
]


class is_edi_cde_cli_line(models.Model):
    _name = "is.edi.cde.cli.line"
    _description="Lignes EDI commandes client"
    _order = "anomalie desc, edi_cde_cli_id,ref_article_client,date_livraison"

    edi_cde_cli_id        = fields.Many2one('is.edi.cde.cli', 'EDI Commandes Clients', required=True, ondelete='cascade')
    num_commande_client   = fields.Char('N° Cde Client')
    ref_article_client    = fields.Char('Ref Article Client')
    product_id            = fields.Many2one('product.product', 'Article')
    quantite              = fields.Integer('Quantité')
    date_livraison        = fields.Date('Date liv')
    numero_document       = fields.Char('N° Document')          # NumeroDocument
    numero_identification = fields.Char('N° Identification')    # NumeroIdentificationAcheteur
    type_commande         = fields.Selection([('ferme', 'Ferme'),('previsionnel', 'Prév.')], "Type")
    prix                  = fields.Float('Prix', digits=(14,4),)
    order_id              = fields.Many2one('sale.order', 'Cde Odoo')
    anomalie              = fields.Text('Anomalie')
    file_id               = fields.Many2one('ir.attachment', 'Fichier')

    date_heure_livraison_au_plus_tot = fields.Char('Livraison au plus tôt' , help="Champ 'DateHeurelivraisonAuPlusTot' pour EDI Weidplas")
    date_heure_livraison             = fields.Char('Livraison au plus tard', help="Champ 'DateHeurelivraisonAuPlusTard' pour EDI Weidplas")
    code_routage                     = fields.Char('Code routage'          , help="Champ 'CodeRoutage' pour EDI Weidplas")
    point_destination                = fields.Char('Point destination'     , help="Champ 'CodeIdentificationPointDestination' pour EDI Weidplas")
    point_dechargement               = fields.Char('Point de déchargement' , help="Champ 'CodeIdentificationPointDechargement' pour EDI Weidplas")
    numero_document                  = fields.Char('N°Document (CALDEL)'   , help="Champ 'NumeroDocument' pour EDI Weidplas => N°UM de PSA")
    tg_number                        = fields.Char('TG Number'             , help="Champ 'TGNumber' pour EDI Weidplas => N°UM de Weidplas")


    def action_acceder_commande(self):
        view_id = self.env.ref('sale.view_order_form').id
        for obj in self:
            return {
                'name': "Commande",
                'view_mode': 'form',
                'view_id': view_id,
                'res_model': 'sale.order',
                'type': 'ir.actions.act_window',
                'res_id': obj.order_id.id,
                'domain': '[]',
            }


class is_edi_cde_cli(models.Model):
    _name = "is.edi.cde.cli"
    _description = "EDI commandes clients"
    _order = "name desc,partner_id"


    @api.depends('line_ids')
    def _compute(self):
        for obj in self:
            r=self.env['is.edi.cde.cli.line'].search([
                ('edi_cde_cli_id','=',obj.id),
            ])
            obj.nb_lignes=len(r)
            r=self.env['is.edi.cde.cli.line'].search([
                ('edi_cde_cli_id','=',obj.id),
                ('anomalie','!=','')
            ])
            obj.nb_anomalies=len(r)


    @api.depends('partner_id')
    def _compute_nb_file(self):
        for obj in self:
            obj.nb_fichiers = len(obj.file_ids)


    name            = fields.Date('Date de création', readonly='1', default=fields.Datetime.now)
    partner_id      = fields.Many2one('res.partner', 'Client', required=False)
    date_maxi       = fields.Date(u"Date de livraison limite d'intégration", help=u"Au delà de cette date, les nouvelles commandes ne seront pas importés et les commandes existantes ne seront pas supprimées")
    jour_semaine    = fields.Selection(_JOURS_SEMAINE, "Jour semaine"      , help=u"Jour de la semaine d'intégration du prévisionnel")
    date_debut_prev = fields.Date(u"Date de début du prévisionnel"         , help=u"A partir de cette date, toutes les commandes seront forcées en prévisionnel")
    import_function = fields.Selection(related="partner_id.is_import_function")
    file_ids        = fields.Many2many('ir.attachment', 'is_doc_attachment_rel', 'doc_id', 'file_id', 'Fichiers')
    create_id       = fields.Many2one('res.users', 'Importe par', readonly=True)
    create_date     = fields.Datetime("Date d'importation")
    state           = fields.Selection([('analyse', u'Analyse'),('traite', u'Traité')], u"État", index=True, default="analyse")
    line_ids        = fields.One2many('is.edi.cde.cli.line', 'edi_cde_cli_id', u"Commandes a importer")
    nb_lignes       = fields.Integer("Nombre de lignes"  , compute='_compute'          , readonly=True, store=False)
    nb_fichiers     = fields.Integer("Nombre de fichier" , compute='_compute_nb_file'  , readonly=True, store=False)
    nb_anomalies    = fields.Integer("Nombre d'anomalies", compute='_compute'          , readonly=True, store=False)


    def action_analyser_fichiers(self):
        for obj in self:
            for row in obj.line_ids:
                row.unlink()
            line_obj = self.env['is.edi.cde.cli.line']
            for attachment in obj.file_ids:
                datas = self.get_data(obj.import_function, attachment)
                for row in datas:
                    num_commande_client = row["num_commande_client"]
                    ref_article_client  = row["ref_article_client"]
                    point_dechargement = False
                    if "point_dechargement" in row:
                        point_dechargement=row["point_dechargement"]
                    order_id = False
                    date_livraison=False
                    type_commande=False
                    if 'order_id' in row:
                        order_id=row['order_id']
                        order=self.env['sale.order'].search([('id', '=', order_id)])
                    else:
                        order=self.env['sale.order'].search([
                            ('partner_id.is_code', '=', obj.partner_id.is_code),
                            ('is_ref_client'     , '=', ref_article_client),
                            ('client_order_ref'  , '=', num_commande_client),
                            ('is_type_commande'  , '=', 'ouverte'),
                            ('state'             , '=', 'draft'),
                        ])
                    anomalie1   = "Cde non trouvée"
                    pricelist=False
                    if len(order):
                        anomalie1=False
                        order_id     = order[0].id
                        partner_id   = order[0].partner_id.id
                        pricelist    = order[0].pricelist_id
                    for ligne in row["lignes"]:
                        product_id = False
                        prix       = 0
                        anomalie2  = []
                        if len(order):
                            if point_dechargement:
                                if point_dechargement!=order[0].is_point_dechargement:
                                    anomalie2.append(u"Point de déchargement modifié (%s<>%s)"%(point_dechargement,order[0].is_point_dechargement))
                        if "anomalie" in ligne:
                            if ligne["anomalie"]:
                                anomalie2.append(ligne["anomalie"])
                        if len(order):
                            if len(order)>1:
                                t=[]
                                for o in order:
                                    t.append(o.name)
                                anomalie2.append('Commande ouverte en double trouvée =>'+','.join(t))
                                order=order[0]
                            quantite   = int(ligne["quantite"])
                            if 'product' in row:
                                product = row['product']
                            else:
                                product    = order[0].is_article_commande_id
                            product_id = product.id

                            #** Date de livraison sur le jour indiqué **********
                            date_livraison=ligne["date_livraison"]
                            if isinstance(date_livraison, str):
                                if date_livraison!='':
                                    date_livraison=datetime.strptime(ligne["date_livraison"], "%Y-%m-%d").date()
                            if isinstance(date_livraison, datetime):
                                date_livraison=date_livraison.date()
                            if ligne["type_commande"]=='previsionnel' and date_livraison and obj.jour_semaine:
                                d=date_livraison
                                jour_semaine_client = d.weekday() + 1
                                jour_semaine        = int(obj.jour_semaine)
                                delta = jour_semaine - jour_semaine_client
                                if delta:
                                    d = d + timedelta(days=delta)
                                    date_livraison = d
                            #***************************************************


                            #** Recherche du prix ******************************
                            if quantite>0:
                                context={}
                                if pricelist:
                                    ctx = dict(
                                        context,
                                        uom=product.uom_id.id,
                                        date=date_livraison,
                                    )
                                    prix, justifcation = pricelist.price_get(
                                        product = product,
                                        qty     = quantite, 
                                        date    = date_livraison
                                    )
                                if prix==0 and obj.import_function not in ('Valeo'):
                                    anomalie2.append("Prix à 0")
                            #***************************************************

                            #** Vérification que qt >= lot livraison ***********
                            lot=self.env['product.template'].get_lot_livraison(product.product_tmpl_id, order.partner_id)

                            if quantite<lot and quantite>0:
                                anomalie2.append("Quantité < Lot de livraison ("+str(int(lot))+")")
                            #***************************************************

                            #** Vérification mutliple du lot *******************
                            arrondi_lot=self.env['product.template'].get_arrondi_lot_livraison(product, order.partner_id, quantite)
                            if quantite!=arrondi_lot and quantite>0:
                                anomalie2.append("Quantité non multiple du lot ("+str(int(arrondi_lot))+")")
                            #***************************************************

                            #** Vérification de la date de livraison livraison *
                            if date_livraison:
                                check_date = self.env['sale.order.line'].check_date_livraison(date_livraison, order.partner_id)
                                if not check_date:
                                    anomalie2.append("Date de livraison pendant la fermeture du client")
                            else:
                                anomalie2.append("Date de livraison non trouvee")
                                date_livraison=False
                            #***************************************************

                            #** En prévisionnel à partir de date_debut_prev ****
                            type_commande = ligne["type_commande"]
                            if obj.date_debut_prev and date_livraison and date_livraison>=obj.date_debut_prev:
                                type_commande = 'previsionnel'
                            #***************************************************

                        if anomalie1:
                            anomalie2.append(anomalie1)
                        anomalie=''
                        if len(anomalie2)>0:
                            anomalie='\n'.join(anomalie2)


                        if obj.import_function=="STELLANTIS":
                            point_dechargement = ligne.get('point_dechargement')

                        vals={
                            'edi_cde_cli_id'       : obj.id,
                            'num_commande_client'  : num_commande_client,
                            'ref_article_client'   : ref_article_client,
                            'product_id'           : product_id,
                            'date_livraison'       : date_livraison,
                            'point_dechargement'   : point_dechargement,
                            'type_commande'        : type_commande,
                            'prix'                 : prix,
                            'order_id'             : order_id,
                            'anomalie'             : anomalie,
                            'file_id'              : attachment.id,
                            'quantite'                         : ligne["quantite"],
                            'numero_document'                  : ligne.get('numero_document'),
                            'tg_number'                        : ligne.get('tg_number'),
                            'numero_identification'            : ligne.get('numero_identification'),
                            'date_heure_livraison'             : ligne.get('date_heure_livraison'),
                            'date_heure_livraison_au_plus_tot' : ligne.get('date_heure_livraison_au_plus_tot'),
                            'code_routage'                     : ligne.get('code_routage'),
                            'point_destination'                : ligne.get('point_destination'),
                        }
                        line_obj.create(vals)


    def action_detail_lignes(self):
        for obj in self:
            if obj.import_function=="STELLANTIS":
                view_id=self.env.ref('is_plastigray16.is_edi_cde_cli_line_tree_view_weidplast')
            else:
                view_id=self.env.ref('is_plastigray16.is_edi_cde_cli_line_tree_view')
            return {
                'name'     : 'Lignes %s'%obj.name,
                'view_mode': 'tree',
                'view_id'  : view_id.id,
                'res_model': 'is.edi.cde.cli.line',
                'context'  : {"search_default_anomalies":1},
                'type'     : 'ir.actions.act_window',
                'domain'   : [
                    ('edi_cde_cli_id','=',obj.id),
                ],
            }


    def action_importer_commandes(self):
        for obj in self:
            line_obj       = self.env['sale.order.line']

            #** Pour PK et Watts, il faut supprimer toutes les commandes de tous les articles 
            if obj.import_function=="Plasti-ka":
                filtre=[
                    ('is_type_commande'  , '=', 'ouverte'),
                    ('state'             , '=', 'draft'),
                    ('partner_invoice_id', '=', obj.partner_id.id),
                ]
                orders=self.env['sale.order'].search(filtre)
            if obj.import_function in ["Watts","SIMU-SOMFY"]:
                filtre=[
                    ('is_type_commande', '=', 'ouverte'),
                    ('state'           , '=', 'draft'),
                    ('partner_id'      , '=', obj.partner_id.id),
                ]
                orders=self.env['sale.order'].search(filtre)
            if obj.import_function in ["Plasti-ka", "Watts", "SIMU-SOMFY"]:
                for order in orders:
                    filtre=[
                        ('order_id'        , '=', order.id),
                        ('is_type_commande', '=', 'previsionnel'),
                    ]
                    line_obj.search(filtre).unlink()
            #*******************************************************************


            #** Recherche des commandes ouvertes trouvées **********************
            order_ids={}
            for line in obj.line_ids:
                if line.order_id:
                    order_ids[line.order_id.id]=True
            #*******************************************************************

            #** Suppression des anciennes commandes ****************************
            #date_jour=time.strftime('%Y-%m-%d')
            date_jour = datetime.now()
            for order_id in order_ids:
                filtre=[
                    ('order_id', '=', order_id),
                    ('is_date_livraison', '>=', date_jour),
                ]
                #Ne pas supprimer les commandes fermes
                if obj.import_function=="eCar":
                    filtre.append(('is_type_commande', '!=', 'ferme'))
                #Ne pas supprimer les commandes au dela de la date limite
                if obj.date_maxi:
                    filtre.append(('is_date_livraison', '<=', obj.date_maxi))
                #Pour plasti-ka, supprimer toutes les commandes => Supprimer toutes les commandes de tous les articles
                #if obj.import_function=="Plasti-ka":
                #    filtre=[
                #        ('order_id', '=', order_id),
                #    ]
                order_line=line_obj.search(filtre)
                for row in order_line:
                    row.unlink()
            #*******************************************************************


            #** Importation des commandes **************************************
            sequence=0
            lines=self.env['is.edi.cde.cli.line'].search([('edi_cde_cli_id','=',obj.id)],order='edi_cde_cli_id,ref_article_client,date_livraison')
            orders=[]

            #pr=cProfile.Profile()
            #pr.enable()
            for line in lines:
                if line.order_id:

                    #** Pour THERMOR supprimer la ligne avant de la créer car ils envoient des commandes inférieures à la date du jour 
                    if obj.import_function=='THERMOR':
                        date_jour=time.strftime('%Y-%m-%d')
                        filtre=[
                            ('order_id', '=', line.order_id.id),
                            ('is_date_livraison', '=', line.date_livraison),
                            ('is_type_commande','=',line.type_commande),
                        ]
                        lines=line_obj.search(filtre)
                        lines.unlink()
                    #***************************************************

                    if line.quantite!=0 and order_id:
                        #Ne pas importer les commandes au dela de la date limite
                        test=True
                        if obj.date_maxi:
                            if line.date_livraison>obj.date_maxi:
                                test=False
                        if test:
                            order=line.order_id
                            if order not in orders:
                                orders.append(order)


                            sequence=sequence+10
                            vals={
                                'sequence'            : sequence,
                                'order_id'            : line.order_id.id, 
                                'is_date_livraison'   : line.date_livraison, 
                                'is_type_commande'    : line.type_commande, 
                                'product_id'          : line.order_id.is_article_commande_id.id or line.product_id.id, 
                                'product_uom_qty'     : line.quantite, 
                                'is_client_order_ref' : line.order_id.client_order_ref, 
                                'price_unit'          : line.prix,
                                'is_date_heure_livraison_au_plus_tot': line.date_heure_livraison_au_plus_tot,
                                'is_numero_document'                 : line.numero_document,
                                'is_tg_number'                       : line.tg_number,
                                'is_code_routage'                    : line.code_routage,
                                'is_point_destination'               : line.point_destination,
                            }
                            line_obj.create(vals)
            #pr.disable()
            #pr.dump_stats('/tmp/action_importer_commandes.cProfile')
            #*******************************************************************

            #** Numérotation des lignes des commandes **************************
            for order in orders:
                order.numeroter_lignes() 
            #*******************************************************************
            obj.state='traite'


    def group_by_data(self, datas):
        dict={}
        for data in datas:
            ref_article_client  = data.get("ref_article_client")
            num_commande_client = data.get("num_commande_client")
            point_dechargement  = data.get("point_dechargement")
            order_id            = data.get("order_id")
            product             = data.get("product")
            lignes              = data.get("lignes",[])
            key="%s-%s-%s-%s-%s"%(ref_article_client,num_commande_client,point_dechargement,order_id,product)
            if key not in dict:
                vals={
                    "ref_article_client" : ref_article_client,
                    "num_commande_client": num_commande_client,
                    "lignes"             : [],
                }
                if order_id:
                    vals["order_id"] = order_id
                if point_dechargement:
                    vals["point_dechargement"] = point_dechargement
                if product:
                    vals["product"] = product
                dict[key]={}
                dict[key]["vals"]=vals
            dict[key]["vals"]["lignes"]+=lignes
        for key in dict:
            dict_ligne={}
            for ligne in dict[key]["vals"]["lignes"]:
                type_commande  = ligne.get("type_commande")
                anomalie       = ligne.get("anomalie")
                date_livraison = ligne.get("date_livraison")
                quantite       = ligne.get("quantite")
                key_ligne="%s-%s-%s"%(type_commande,anomalie,date_livraison)
                if key_ligne not in dict_ligne:
                    dict_ligne[key_ligne]={}
                    dict_ligne[key_ligne]["type_commande"]=type_commande
                    dict_ligne[key_ligne]["anomalie"]=anomalie
                    dict_ligne[key_ligne]["date_livraison"]=date_livraison
                    dict_ligne[key_ligne]["quantite"]=0
                dict_ligne[key_ligne]["quantite"]+=quantite
            dict[key]["vals"]["lignes"]=[]
            for key_ligne in dict_ligne:
                dict[key]["vals"]["lignes"].append(dict_ligne[key_ligne])
        datas=[]
        for key in dict:
            datas.append(dict[key]["vals"])
        return datas


    def get_data(self, import_function, attachment):
        datas={}

        if import_function=="902580":
            datas=self.get_data_902580(attachment)
        if import_function=="902810":
            datas=self.get_data_902810(attachment)
        if import_function=="903410":
            datas=self.get_data_903410(attachment)
        if import_function=="ACTIA":
            datas=self.get_data_ACTIA(attachment)
        if import_function=="ASTEELFLASH":
            datas=self.get_data_ASTEELFLASH(attachment)
        if import_function=="DARWIN":
            datas=self.get_data_DARWIN(attachment)
        if import_function=="eCar":
            datas=self.get_data_eCar(attachment)
            datas=self.group_by_data(datas)
        if import_function=="GXS":
            datas=self.get_data_GXS(attachment)
        if import_function=="John-Deere":
            datas=self.get_data_John_Deere(attachment)
        if import_function=="Millipore":
            datas=self.get_data_Millipore(attachment)
        if import_function=="Mini-Delta-Dore":
            datas=self.get_data_MiniDeltaDore(attachment)
        if import_function=="Motus":
            datas=self.get_data_Motus(attachment)
        if import_function == 'Lacroix':
            datas = self.get_data_lacroix(attachment)
        if import_function=="Odoo":
            datas=self.get_data_Odoo(attachment)
        if import_function=="Plasti-ka":
            datas=self.get_data_plastika(attachment)
        if import_function=="SIMU":
            datas=self.get_data_SIMU(attachment)
        if import_function=="SIMU-SOMFY":
            datas=self.get_data_SIMU_SOMFY(attachment)
        if import_function=="STELLANTIS":
            datas=self.get_data_STELLANTIS(attachment)
        if import_function=="THERMOR":
            datas=self.get_data_THERMOR(attachment)
        if import_function=="Valeo":
            datas=self.get_data_Valeo(attachment)
        if import_function=="Watts":
            datas=self.get_data_Watts(attachment)

        return datas


    def getNumCommandeClient(self, ref_article_client):
        """Recherche du numéro de commande client à partir de la référence article"""
        for obj in self:
            num_commande_client = "??"
            SaleOrder = self.getSaleOrder(ref_article_client)
            if SaleOrder:
                num_commande_client = SaleOrder.client_order_ref
            return num_commande_client


    def getSaleOrder(self, ref_article_client):
        """Recherche de la commande ouverte client à partir de la référence article"""
        for obj in self:
            order = self.env['sale.order'].search([
                ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                ('is_ref_client', '=', ref_article_client),
                ('is_type_commande'  , '=', 'ouverte'),
            ])
            SaleOrder = False
            if len(order):
                SaleOrder = order[0]
            return SaleOrder





    def get_data_ASTEELFLASH(self, attachment):
        res = []
        for obj in self:
            #** Lecture du fichier xlsx ****************************************
            xlsxfile = base64.decodebytes(attachment.datas)
            path = '/tmp/edi-asteelflash-'+str(obj.id)+'.xlsx'
            with open(path,'wb') as f:
                f.write(xlsxfile)
            #*******************************************************************

            #** Test si fichier est bien du xlsx *******************************
            try:
                wb = openpyxl.load_workbook(filename = path)
                ws = wb.active
                title = ws.title
                cells = list(ws)
            except:
                raise ValidationError(u"Le fichier "+attachment.name+u" n'est pas un fichier xlsx")
            #*******************************************************************

            lig=0
            for row in ws.rows:
                if lig>0:
                    type_commande="previsionnel"
                    ref_article_client = cells[lig][0].value
                    try:
                        quantite = cells[lig][4].value
                        quantite=float(quantite)
                    except:
                        quantite=0
                    date_livraison = cells[lig][6].value
                    try:
                        date_livraison = date_livraison.strftime('%Y-%m-%d')
                    except ValueError:
                        date_livraison = False
                    if date_livraison:
                        order = self.env['sale.order'].search([
                            ('partner_id.is_code', '=', obj.partner_id.is_code),
                            ('is_ref_client'     , '=', ref_article_client),
                            ('is_type_commande'  , '=', 'ouverte')]
                        )
                        num_commande_client = "??"
                        if len(order):
                            num_commande_client = order[0].client_order_ref
                        val = {
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        ligne = {
                            'quantite'      : quantite,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes': [ligne]})
                        res.append(val)
                lig+=1
        return res


    def get_data_SIMU(self, attachment):
        res = []
        for obj in self:
            #** Lecture du fichier xlsx ****************************************
            xlsxfile = base64.decodebytes(attachment.datas)
            path = '/tmp/edi-simu-'+str(obj.id)+'.xlsx'
            f = open(path,'wb')
            f.write(xlsxfile)
            f.close()
            type_fichier=False
            #*******************************************************************

            #** Test si fichier contenant le prévisionnel **********************
            try:
                type_fichier="previsionnel"
                wb = openpyxl.load_workbook(filename = path)
                ws = wb['DL']
                cells = list(ws)
            except:
                type_fichier=False
            #*******************************************************************

            #** Test si fichier contenant le ferme *****************************
            if type_fichier==False:
                try:
                    type_fichier="ferme"
                    wb = openpyxl.load_workbook(filename = path)
                    ws = wb[u'OA non réceptionnés']
                    cells = list(ws)
                except:
                    type_fichier=False
            #*******************************************************************

            if type_fichier=='ferme':
                lig=0
                for row in ws.rows:
                    if lig>0:
                        type_commande="ferme"
                        try:
                            quantite = cells[lig][15].value
                            quantite=float(quantite)
                        except ValueError:
                            quantite=0
                        ref_article_client = cells[lig][2].value

                        #** Recherche article **********************************
                        products = self.env['product.product'].search([
                            ('is_client_id.is_code', '=', obj.partner_id.is_code),
                            ('is_ref_client'     , '=', ref_article_client),
                        ])
                        Product = False
                        if len(products):
                            Product = products[0]
                        #*******************************************************


                        SaleOrder = False
                        num_commande_client = "??"
                        if Product:
                            #** Recherche et création de la commande ferme *****
                            num_commande_client = cells[lig][9].value
                            order = self.env['sale.order'].search([
                                ('partner_id.is_code', '=', obj.partner_id.is_code),
                                #('is_ref_client'     , '=', ref_article_client),
                                ('is_type_commande'  , '=', 'standard'),
                                ('client_order_ref'  , '=', num_commande_client),
                            ])

                            if len(order):
                                SaleOrder = order[0]
                            else:
                                vals={
                                    'partner_id'      : obj.partner_id.id,
                                    'is_type_commande': 'standard',
                                    'client_order_ref': num_commande_client,
                                    'pricelist_id'    : obj.partner_id.property_product_pricelist.id,
                                }
                                SaleOrder = self.env['sale.order'].create(vals)
                                SaleOrder.pg_onchange_partner_id()
                            #***************************************************


                        #** Répartir la quantité sur 4 jours *******************
                        lot_livraison = 0
                        if SaleOrder:
                            #num_commande_client = SaleOrder.client_order_ref
                            for line in Product.is_client_ids:
                                if line.client_id == obj.partner_id:
                                    lot_livraison = line.lot_livraison
                                    nb_lots = math.ceil(quantite / lot_livraison / 4.0)
                                    lot_livraison = lot_livraison * nb_lots
                        quantites=[]
                        if quantite>lot_livraison and lot_livraison>0:
                            nb_lots = quantite / lot_livraison / 4.0
                            reste = quantite
                            for x in range(4):
                                v = lot_livraison
                                if (reste-v)<0:
                                    v = reste
                                reste = reste - v
                                quantites.append(v)
                                if reste==0:
                                    break
                        else:
                            quantites.append(quantite)
                        #*******************************************************

                        #** Mettre la date au lundi ****************************
                        try:
                            date_livraison = cells[lig][12].value
                            jour_semaine = date_livraison.weekday() + 7 # Semaine précédente
                            date_lundi = date_livraison - timedelta(days=jour_semaine)
                        except ValueError:
                            date_livraison = False
                            jour_semaine   = False
                            date_lundi     = False
                        #*******************************************************

                        for quantite in quantites:
                            date_livraison = False
                            if date_lundi:
                                date_livraison = date_lundi.strftime('%Y-%m-%d')
                                date_lundi = date_lundi + timedelta(days=1)
                            val = {
                                'num_commande_client' : num_commande_client,
                                'ref_article_client'  : ref_article_client,
                                'order_id'            : SaleOrder and SaleOrder.id,
                                'product'             : Product,
                            }
                            ligne = {
                                'quantite'      : quantite,
                                'type_commande' : type_commande,
                                'date_livraison': date_livraison,
                            }
                            val.update({'lignes': [ligne]})
                            res.append(val)
                    lig+=1

            if type_fichier=='previsionnel':
                now = datetime.now()
                annee = int(now.year)
                lig=0
                semaines={}
                test=False
                mem_ref = ''
                for row in ws.rows:
                    nb_cols = len(row)
                    if nb_cols>100:
                        nb_cols=100

                    #** Traitement des lignes des prévisions *******************
                    if test:
                        val1=cells[lig][0].value
                        val2=cells[lig][1].value
                        if val1:
                            mem_ref=val1
                        if val2==u'Forecast / prévision':
                            vals={}
                            for col in range(nb_cols):
                                if col>=2:
                                    v = cells[lig][col].value
                                    if v:
                                        semaine = semaines[col+1]
                                        ref_article_client  = mem_ref
                                        num_commande_client = self.getNumCommandeClient(ref_article_client)
                                        #** Convertir la semaine en date *******
                                        sem=''
                                        try:
                                            annee = int(semaine[0:4])
                                            sem   = int(semaine[6:8])
                                            date_livraison = datetime.strptime('%04d-%02d-1' % (annee, sem), '%Y-%W-%w')
                                            #Pour avoir la semaine en ISO car pas dispo en Python 2.7, uniqument avec Python 3
                                            if date(annee, 1, 4).isoweekday() > 4:
                                                date_livraison -= timedelta(days=7)
                                            date_livraison = date_livraison.strftime('%Y-%m-%d')
                                        except:
                                            date_livraison = False
                                        #*******************************************

                                        try:
                                            qt=float(v)
                                        except ValueError:
                                            qt=0

                                        type_commande="previsionnel"
                                        val = {
                                            'num_commande_client' : num_commande_client,
                                            'ref_article_client'  : ref_article_client,
                                        }
                                        ligne = {
                                            'quantite'      : qt,
                                            'type_commande' : type_commande,
                                            'date_livraison': date_livraison,
                                        }
                                        val.update({'lignes': [ligne]})
                                        res.append(val)
                    #***************************************************************


                    #** Recherche des numéros des semaines *************************
                    #nbcols = len(row)-1
                    if cells[lig][1].value=='Semaine / week':
                        for col in range(nb_cols):
                            if col>=2 and cells[lig][col].value:
                                semaine = int(cells[lig][col].value)
                                if col==2:
                                    memsemaine=semaine
                                if semaine<(memsemaine-5):
                                    annee+=1
                                val = cells[lig][col].value
                                semaines[col+1] = str(annee)+'-S'+str(val)
                                memsemaine=semaine
                        test = True # Il est possible de traiter les lignes suivantes
                    #***************************************************************
                    lig+=1
            if type_fichier==False:
                raise ValidationError(u"Le fichier "+attachment.name+u" n'est pas un fichier SIMU compatible (xlsx pour le ferme ou le prévisionnel)")
        return res


    def get_data_SIMU_SOMFY(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode()
            csvfile = csvfile.split("\n")
            csvfile = csv.reader(csvfile, delimiter=';')
            for ct, lig in enumerate(csvfile):
                if ct>0 and len(lig)>=28:
                    #type_commande      = lig[25].strip()
                    type_commande      = lig[7].strip()

                    if type_commande == u"Prévision":
                        type_commande="previsionnel"

                        #ref_article_client = lig[18].strip()
                        ref_article_client = lig[0].strip()

                        #quantite = lig[26].replace(',', '.')
                        quantite = lig[8].replace(',', '.')
                        try:
                            quantite = float(quantite)
                        except ValueError:
                            quantite=0

                        #date_livraison = lig[28]
                        date_livraison = lig[12]

                        d=False
                        try:
                            d = datetime.strptime(date_livraison, '%d/%m/%Y')
                        except ValueError:
                            continue
                        if d:
                            date_livraison = d.strftime('%Y-%m-%d')
                            order = self.env['sale.order'].search([
                                ('partner_id.is_code', '=', obj.partner_id.is_code),
                                ('is_ref_client'     , '=', ref_article_client),
                                ('is_type_commande'  , '=', 'ouverte')]
                            )
                            num_commande_client = "??"
                            if len(order):
                                num_commande_client = order[0].client_order_ref
                            val = {
                                'num_commande_client' : num_commande_client,
                                'ref_article_client'  : ref_article_client,
                            }
                            ligne = {
                                'quantite'      : quantite,
                                'type_commande' : type_commande,
                                'date_livraison': date_livraison,
                            }
                            val.update({'lignes': [ligne]})
                            res.append(val)
        return res



    def get_data_Millipore(self, attachment):
        res = []
        #mois=[u'janv.',u'févr.',u'mars',u'avr.',u'mai',u'juin',u'juil.',u'août',u'sept.',u'oct.',u'nov.',u'déc.']
        mois=['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode('cp1252')
            csvfile = csvfile.split("\r\n")
            csvfile = csv.reader(csvfile, delimiter='\t')
            tab=[]
            annees=[]
            dates=[]
            for ct, lig in enumerate(csvfile):
                if ct<7:
                    continue
                nb=len(lig)
                if ct==7:
                    for i in range(7,nb):
                        annees.append(lig[i])
                if ct==8:
                    for i in range(7,nb):
                        m=mois.index(lig[i])+1
                        txt=str(annees[i-7])+'-'+str(m)+'-01'
                        d=datetime.strptime(txt, '%Y-%m-%d')
                        #** Recherche du premier mercredi du mois **************
                        while True:
                            jour_semaine=d.strftime('%w')
                            if jour_semaine=='3':
                                break
                            d = d + timedelta(days=1)
                        #*******************************************************
                        dates.append(d.strftime('%Y-%m-%d'))
                if ct>8:
                    if nb>=10:
                        ref_article_client = lig[0].strip()
                        order = self.env['sale.order'].search([
                            ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                            ('is_ref_client', '=', ref_article_client)]
                        )
                        num_commande_client = "??"
                        if len(order):
                            num_commande_client = order[0].client_order_ref
                        for i in range(7,nb):
                            val = {
                                'num_commande_client' : num_commande_client,
                                'ref_article_client'  : ref_article_client,
                            }
                            date_livraison=dates[(i-7)]
                            quantite = lig[i]
                            try:
                                qt = float(quantite)
                            except ValueError:
                                qt=0
                            type_commande="previsionnel"
                            ligne = {
                                'quantite'      : qt,
                                'type_commande' : type_commande,
                                'date_livraison': date_livraison,
                            }
                            val.update({'lignes': [ligne]})
                            res.append(val)
        return res


    def get_data_MiniDeltaDore(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode()
            csvfile = csvfile.split("\n")
            csvfile = csv.reader(csvfile, delimiter='\t')
            for ct, lig in enumerate(csvfile):
                if len(lig)==4:
                    ref_article_client = lig[0].strip()
                    order = self.env['sale.order'].search([
                        ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                        ('is_ref_client', '=', ref_article_client),
                        ('is_type_commande'  , '=', 'ouverte'),
                    ])
                    num_commande_client = "??"
                    if len(order):
                        num_commande_client = order[0].client_order_ref

                    val = {
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                    }
                    date_livraison=lig[1]
                    quantite = lig[3]
                    try:
                        quantite = float(quantite)
                    except ValueError:
                        quantite=0
                    type_commande=lig[2]
                    ligne = {
                        'quantite'      : quantite,
                        'type_commande' : type_commande,
                        'date_livraison': date_livraison,
                    }
                    val.update({'lignes': [ligne]})
                    res.append(val)
        return res






    def get_data_ACTIA(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode()
            csvfile = csvfile.split("\n")
            csvfile = csv.reader(csvfile, delimiter=';')
            tab=[]
            for ct, lig in enumerate(csvfile):
                if ct == 0:
                    continue
                if len(lig) == 9:
                    ref_article_client = lig[1].strip()
                    order = self.env['sale.order'].search([
                        ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                        ('is_ref_client', '=', ref_article_client)]
                    )
                    num_commande_client = "??"
                    if len(order):
                        num_commande_client = order[0].client_order_ref
                    val = {
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                    }
                    quantite = lig[4]
                    qt=0
                    try:
                        qt = float(quantite)
                    except ValueError:
                        continue
                    type_commande="previsionnel"
                    date_livraison = lig[3].strip()
                    d=False
                    try:
                        d = datetime.strptime(date_livraison, '%Y%m%d')
                    except ValueError:
                        continue
                    if d:
                        date_livraison = d.strftime('%Y-%m-%d')
                        ligne = {
                            'quantite'      : qt,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes': [ligne]})
                        res.append(val)
        return res






    def get_data_DARWIN(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode()
            csvfile = csvfile.split("\n")
            csvfile = csv.reader(csvfile, delimiter=',')
            for ct, lig in enumerate(csvfile):
                if ct>0 and len(lig)>=15:
                    ref_article_client  = lig[2].strip()
                    num_commande_client = lig[7].strip()
                    try:
                        date_livraison=lig[10][0:10]
                        d=datetime.strptime(date_livraison, '%d/%m/%Y')
                        date_livraison=d.strftime('%Y-%m-%d')
                    except ValueError:
                        date_livraison=False
                    try:
                        qt=float(lig[4])
                    except ValueError:
                        qt=0
                    type_commande="previsionnel"
                    if lig[15]=='Firm':
                        type_commande='ferme'
                    val = {
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                    }
                    ligne = {
                        'quantite'      : qt,
                        'type_commande' : type_commande,
                        'date_livraison': date_livraison,
                    }
                    val.update({'lignes': [ligne]})
                    res.append(val)
        return res


    def get_data_lacroix(self, attachment):
        nb_cols = 22
        col_ref = 1
        col_qn = 16
        col_type = None
        col_date = 19
        data_previsionnel = 'P'
        res = []
        for obj in self:
            #** Lecture du fichier xlsx ****************************************
            xlsxfile = base64.decodebytes(attachment.datas)
            path = '/tmp/THERMOR-'+str(obj.id)+'.xlsx'
            f = open(path,'wb')
            f.write(xlsxfile)
            f.close()
            type_fichier=False
            #*******************************************************************

            #** Ouverture du fichier *******************************************
            try:
                wb = openpyxl.load_workbook(filename = path)
                ws = wb.active
                cells = list(ws)
            except:
                raise ValidationError(u"Le fichier "+attachment.name+u" n'est pas un fichier xlsx valide")
            try:
                for ct, lig in enumerate(ws):
                    if not ct:
                        continue
                    if len(lig) == nb_cols:
                        ref_article_client = lig[col_ref].value.strip()
                        order = self.env['sale.order'].search([
                            ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                            ('is_ref_client', '=', ref_article_client)]
                        )
                        num_commande_client = "??"
                        if len(order):
                            num_commande_client = order[0].client_order_ref
                        val = {
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        # la quantite est normalement un int
                        quantite = lig[col_qn].value
                        try:
                            qt = float(quantite)
                        except ValueError:
                            continue
                        if col_type is None:
                            type_commande="previsionnel"
                        else:
                            type_commande = lig[col_type].value.strip()
                            if type_commande == data_previsionnel:
                                type_commande = "previsionnel"
                            else:
                                type_commande = "ferme"
                        # La date est un "datetime"
                        date_livraison = lig[col_date].value
                        try:
                            date_livraison = date_livraison.strftime('%Y-%m-%d')
                        except ValueError:
                            date_livraison=False

                        ligne = {
                            'quantite'      : qt,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes': [ligne]})
                        res.append(val)
            except csv.Error:
                raise ValidationError('Fichier vide ou non compatible (le fichier doit être au format CSV)')
        return res



    def get_data_902810(self, attachment):
        res = []
        for obj in self:
            csvfile=base64.decodebytes(attachment.datas)
            csvfile = csvfile.decode("Windows-1252")
            csvfile=csvfile.split("\n")
            tab=[]
            ct=0
            for row in csvfile:
                lig=row.split(";")
                if len(lig)==9:
                    ct=ct+1
                    if ct>1:
                        ref_article_client=lig[0].strip()
                        order=self.env['sale.order'].search([
                            ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                            ('is_ref_client', '=', ref_article_client)]
                        )
                        num_commande_client="??"
                        if len(order):
                            num_commande_client=order[0].client_order_ref
                        val={
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        quantite=lig[6].strip()
                        qt=0
                        try:
                            qt=float(quantite)
                        except ValueError:
                            continue
                        type_commande=lig[5].strip()
                        if type_commande=="P":
                            type_commande="previsionnel"
                        else:
                            type_commande="ferme"
                        date_livraison=lig[4].strip()
                        d=datetime.strptime(date_livraison, '%d/%m/%y')
                        date_livraison=d.strftime('%Y-%m-%d')
                        ligne = {
                            'quantite'      : qt,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes':[ligne]})
                        res.append(val)
        return res


    def get_data_903410(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas)
            csvfile = csvfile.decode("utf-16")
            csvfile = csvfile.split("\r")
            csvfile = csv.reader(csvfile, delimiter='\t')
            for ct, lig in enumerate(csvfile):
                if len(lig)>=10 and ct>0:
                    ref_article_client = lig[5].replace("'","").strip()
                    order = self.env['sale.order'].search([
                        ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                        ('is_ref_client', '=', ref_article_client),
                        ('is_type_commande'  , '=', 'ouverte'),
                    ])
                    num_commande_client = "??"
                    if len(order):
                        num_commande_client = order[0].client_order_ref
                    val = {
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                    }
                    try:
                        date_livraison=lig[9][0:10]
                        d=datetime.strptime(date_livraison, '%d.%m.%Y')
                        date_livraison=d.strftime('%Y-%m-%d')
                    except ValueError:
                        date_livraison=False
                    quantite = lig[10].replace(" ","").replace(",",".").strip()
                    try:
                        quantite = float(quantite)
                    except ValueError:
                        quantite=0
                    type_commande="previsionnel"
                    if lig[7].strip()==u'Finished Material':
                        type_commande="ferme"
                    ligne = {
                        'quantite'      : quantite,
                        'type_commande' : type_commande,
                        'date_livraison': date_livraison,
                    }
                    val.update({'lignes': [ligne]})
                    res.append(val)
        return res




    def get_data_eCar(self, attachment):
        res = []
        for obj in self:
            filename = '/tmp/%s.xml' % attachment.id
            temp = open(filename, 'w+b')
            temp.write((base64.decodebytes(attachment.datas))) 
            temp.close()
            tree = etree.parse(filename)
            for partie_citee in tree.xpath("/DELINS/PARTIE_CITEE"):
                for article_programme in partie_citee.xpath("ARTICLE_PROGRAMME"):
                    ref_article_client=""
                    for NumeroArticleClient in article_programme.xpath("NumeroArticleClient"):
                        ref_article_client=NumeroArticleClient.text
                    num_commande_client=""
                    for NumeroArticleClient in article_programme.xpath("NumeroCommande"):
                        num_commande_client=NumeroArticleClient.text
                    point_dechargement=False
                    for POINT_DE_DECHARGEMENT in article_programme.xpath("POINT_DE_DECHARGEMENT"):
                        for CodeIdentificationPointDechargement in POINT_DE_DECHARGEMENT.xpath("CodeIdentificationPointDechargement"):
                            point_dechargement = CodeIdentificationPointDechargement.text
                    val = {
                        'ref_article_client' : ref_article_client,
                        'num_commande_client': num_commande_client,
                        'point_dechargement' : point_dechargement,
                        'lignes': []
                    }
                    res1 = []
                    for detail_programme in article_programme.xpath("DETAIL_PROGRAMME"):
                        type_commande=""
                        for CodeStatutProgramme in detail_programme.xpath("CodeStatutProgramme"):
                            type_commande=CodeStatutProgramme.text
                        if type_commande=="4":
                            type_commande="previsionnel"
                        else:
                            type_commande="ferme"
                        date_livraison=""
                        for DateHeureLivraisonAuPlusTot in detail_programme.xpath("DateHeureLivraisonAuPlusTot"):
                            date_livraison=DateHeureLivraisonAuPlusTot.text[:8]
                        if date_livraison!="":
                            date_livraison=datetime.strptime(date_livraison, '%Y%m%d')
                        quantite=""
                        for QuantiteALivrer in detail_programme.xpath("QuantiteALivrer"):
                            quantite=QuantiteALivrer.text
                            try:
                                quantite=float(quantite)
                            except ValueError:
                                quantite=0

                        ligne = {
                            'quantite'      : quantite,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        res1.append(ligne)
                    val.update({'lignes':res1})
                    res.append(val)


            for partie_citee in tree.xpath("/CALDEL/SEQUENCE_PRODUCTION"):
                if  partie_citee.xpath("ARTICLE_PROGRAMME"):
                    ref_article_client=partie_citee.xpath("ARTICLE_PROGRAMME/NumeroArticleClient")[0].text
                    products=self.env['product.product'].search([
                        ('is_client_id' , '=', obj.partner_id.id),
                        ('is_ref_client', '=', ref_article_client),
                        ('is_gestionnaire_id', '!=', '04'), 
                        ('is_gestionnaire_id', '!=', '07'), 
                        ('is_gestionnaire_id', '!=', '12'), 
                        ('is_gestionnaire_id', '!=', '14'), 
                        ('is_gestionnaire_id', '!=', '23'), 
                        ('segment_id', 'not ilike', 'fictif'), 
                        ('segment_id', 'not ilike', 'fantome'), 
                        ('segment_id', 'not ilike', 'consommable'), 
                        ('segment_id', 'not ilike', 'comptable'),
                    ])
                    anomalie=False
                    if len(products)==0:
                        anomalie=u"Article '%s' non trouvé pour le client %s/%s"%(ref_article_client,obj.partner_id.is_code,obj.partner_id.is_adr_code)
                    if len(products)>1:
                        anomalie=u"Il existe plusieurs articles actifs pour la référence client '%s' et pour le client %s/%s"%(ref_article_client,obj.partner_id.is_code,obj.partner_id.is_adr_code)
                    num_commande_client=partie_citee.xpath("ARTICLE_PROGRAMME/NumeroCommande")[0].text
                    order_id=False
                    product=False
                    if not anomalie:
                        product=products[0]
                        orders=self.env['sale.order'].search([
                            ('is_type_commande'  , '=', 'standard'),
                            ('state'             , '=', 'draft'),
                            ('partner_id'        , '=', obj.partner_id.id),
                            ('client_order_ref'  , '=', num_commande_client),
                        ])
                        if len(orders)==0:
                            vals={
                                "partner_id": obj.partner_id.id,
                                "client_order_ref": num_commande_client,
                                "is_type_commande": "standard",
                            }
                            order=self.env['sale.order'].create(vals)
                            order.pg_onchange_partner_id()
                        else:
                            order=orders[0]
                        order_id=order.id
                    val = {
                        'order_id'           : order_id,
                        'ref_article_client' : ref_article_client,
                        'num_commande_client': num_commande_client,
                        'lignes'             : []
                    }
                    if product:
                        val["product"]=product
                    res1 = []
                    for detail_programme in partie_citee.xpath("ARTICLE_PROGRAMME/DETAIL_PROGRAMME_ARTICLE"):
                        date_livraison=detail_programme.xpath("DateHeurelivraisonAuPlusTard")[0].text[:8]
                        quantite=detail_programme.xpath("QteALivrer")[0].text
                        try:
                            quantite=float(quantite)
                        except ValueError:
                            quantite=0
                        date_livraison=datetime.strptime(date_livraison, '%Y%m%d')
                        ligne = {
                            'quantite': quantite,
                            'type_commande': 'ferme',
                            'date_livraison': date_livraison,
                            'anomalie'           : anomalie,
                        }
                        res1.append(ligne)
                    val.update({'lignes':res1})
                    res.append(val)
                else:
                    continue

        return res



    def get_data_STELLANTIS(self, attachment):
        res = []
        # def get_products(is_client_id=False,is_ref_client=False):
        #     products=self.env['product.product'].search([
        #         ('is_client_id' , '=', is_client_id),
        #         ('is_ref_client', '=', is_ref_client),
        #         ('is_gestionnaire_id', '!=', '04'), 
        #         ('is_gestionnaire_id', '!=', '07'), 
        #         ('is_gestionnaire_id', '!=', '12'), 
        #         ('is_gestionnaire_id', '!=', '14'), 
        #         ('is_gestionnaire_id', '!=', '23'), 
        #         ('segment_id', 'not ilike', 'fictif'), 
        #         ('segment_id', 'not ilike', 'fantome'), 
        #         ('segment_id', 'not ilike', 'consommable'), 
        #         ('segment_id', 'not ilike', 'comptable'),
        #     ])
        #     return products
        

        for obj in self:
            filename = '/tmp/%s.xml' % attachment.id
            temp = open(filename, 'w+b')
            temp.write((base64.decodebytes(attachment.datas))) 
            temp.close()
            tree = etree.parse(filename)
            for partie_citee in tree.xpath("/DELINS/PARTIE_CITEE"):
                for article_programme in partie_citee.xpath("ARTICLE_PROGRAMME"):
                    ref_article_client=""
                    for NumeroArticleClient in article_programme.xpath("NumeroArticleClient"):
                        ref_article_client=NumeroArticleClient.text
                    num_commande_client=""
                    for NumeroArticleClient in article_programme.xpath("NumeroCommande"):
                        num_commande_client=NumeroArticleClient.text

                    anomalie=[]

                    # products = get_products(is_client_id=obj.partner_id.id,is_ref_client=ref_article_client)
                    # if len(products)==0:
                    #     anomalie.append("Article '%s' non trouvé pour le client %s/%s"%(ref_article_client,obj.partner_id.is_code,obj.partner_id.is_adr_code))
                    # if len(products)>1:
                    #     anomalie.append("Il existe plusieurs articles actifs pour la référence client '%s' et pour le client %s/%s"%(ref_article_client,obj.partner_id.is_code,obj.partner_id.is_adr_code))




                    #product=False
                    order_id=False
                    if not len(anomalie):
                        #product=products[0]
                        orders=self.env['sale.order'].search([
                            ('partner_id.is_code', '=' , obj.partner_id.is_code),
                            ('is_ref_client'     , '=' , ref_article_client),
                            ('is_type_commande'  , '=' , 'ouverte'),
                            ('state'             , '=' , 'draft'),
                            ('client_order_ref'  , '!=', 'PREVISIONS'),
                        ])


                        for order in orders:
                            order_id = order.id
                            if order.client_order_ref!=num_commande_client:
                                anomalie.append("Commande EDI client '%s' différente de la commande Odoo '%s'"%(num_commande_client,order.client_order_ref))
                    point_dechargement=False
                    for POINT_DE_DECHARGEMENT in article_programme.xpath("POINT_DE_DECHARGEMENT"):
                        for CodeIdentificationPointDechargement in POINT_DE_DECHARGEMENT.xpath("CodeIdentificationPointDechargement"):
                            point_dechargement = CodeIdentificationPointDechargement.text
                    val = {
                        'order_id'           : order_id,
                        'ref_article_client' : ref_article_client,
                        'num_commande_client': num_commande_client,
                        'point_dechargement' : point_dechargement,
                        'lignes': []
                    }
                    res1 = []
                    for detail_programme in article_programme.xpath("DETAIL_PROGRAMME"):
                        type_commande=""
                        for CodeStatutProgramme in detail_programme.xpath("CodeStatutProgramme"):
                            type_commande=CodeStatutProgramme.text
                        if type_commande=="4":
                            type_commande="previsionnel"
                        else:
                            type_commande="ferme"

                        date_livraison=""
                        for DateHeureLivraisonAuPlusTot in detail_programme.xpath("DateHeurelivraisonAuPlusTot"):
                            date_livraison=DateHeureLivraisonAuPlusTot.text[:8]

                        if date_livraison=='':
                            for DateHeureLivraisonAuPlusTot in detail_programme.xpath("DateHeureLivraisonAuPlusTot"):
                                date_livraison=DateHeureLivraisonAuPlusTot.text[:8]
                        if date_livraison!="":
                            date_livraison=datetime.strptime(date_livraison, '%Y%m%d')

                        quantite=""
                        for QuantiteALivrer in detail_programme.xpath("QuantiteALivrer"):
                            quantite=QuantiteALivrer.text
                            try:
                                quantite=float(quantite)
                            except ValueError:
                                quantite=0

                        ligne = {
                            'quantite'          : quantite,
                            'type_commande'     : type_commande,
                            'date_livraison'    : date_livraison,
                            'point_dechargement': point_dechargement,
                            'anomalie'          : '\n'.join(anomalie),
                        }
                        res1.append(ligne)
                    val.update({'lignes':res1})
                    res.append(val)

            for caldel in tree.xpath("/CALDEL"):
                numero_document = str(int(caldel.xpath("NumeroDocument")[0].text))
                for partie_citee in caldel.xpath("SEQUENCE_PRODUCTION"):
                    if  partie_citee.xpath("ARTICLE_PROGRAMME"):
                        ref_article_client=partie_citee.xpath("ARTICLE_PROGRAMME/NumeroArticleClient")[0].text

                        try:
                            tg_number=partie_citee.xpath("ARTICLE_PROGRAMME/TGNumber")[0].text
                        except IndexError:
                            tg_number=""

                        #products = get_products(is_client_id=obj.partner_id.id,is_ref_client=ref_article_client)
                        anomalie=[]
                        # if len(products)==0:
                        #     anomalie.append("Article '%s' non trouvé pour le client %s/%s"%(ref_article_client,obj.partner_id.is_code,obj.partner_id.is_adr_code))
                        # if len(products)>1:
                        #     anomalie.append("Il existe plusieurs articles actifs pour la référence client '%s' et pour le client %s/%s"%(ref_article_client,obj.partner_id.is_code,obj.partner_id.is_adr_code))


                        try:
                            num_commande_client=partie_citee.xpath("ARTICLE_PROGRAMME/NumeroCommande")[0].text
                        except IndexError:
                            num_commande_client=""



                        #product=False
                        order_id=False
                        if not len(anomalie):
                            #product=products[0]
                            orders=self.env['sale.order'].search([
                                ('partner_id.is_code', '=', obj.partner_id.is_code),
                                ('is_ref_client'     , '=', ref_article_client),
                                ('is_type_commande'  , '=', 'ouverte'),
                                ('state'             , '=', 'draft'),
                                #('client_order_ref'  , '!=', 'PREVISIONS'),
                            ])
                            for order in orders:
                                order_id = order.id
                                if order.client_order_ref!=num_commande_client:
                                    anomalie.append("Commande EDI client '%s' différente de la commande Odoo '%s'"%(num_commande_client,order.client_order_ref))
                        val = {
                            'order_id'           : order_id,
                            'ref_article_client' : ref_article_client,
                            'num_commande_client': num_commande_client,
                            'lignes'             : []
                        }
                        #if product:
                        #    val["product"]=product
                        point_dechargement = False
                        code_routage       = False
                        point_destination  = False
                        for POINT_DE_DECHARGEMENT in partie_citee.xpath("ARTICLE_PROGRAMME/POINT_DE_DECHARGEMENT"):
                                for CodeIdentificationPointDechargement in POINT_DE_DECHARGEMENT.xpath("CodeIdentificationPointDechargement"):
                                    point_dechargement = CodeIdentificationPointDechargement.text
                                for CodeRoutage in POINT_DE_DECHARGEMENT.xpath("CodeRoutage"):
                                    code_routage = CodeRoutage.text
                                for CodeIdentificationPointDestination in POINT_DE_DECHARGEMENT.xpath("CodeIdentificationPointDestination"):
                                    point_destination = CodeIdentificationPointDestination.text
                        res1 = []
                        for detail_programme in partie_citee.xpath("ARTICLE_PROGRAMME/DETAIL_PROGRAMME_ARTICLE"):
                            quantite                          = detail_programme.xpath("QteALivrer")[0].text
                            date_heure_livraison              = detail_programme.xpath("DateHeurelivraisonAuPlusTard")[0].text
                            date_heure_livraison_au_plus_tot  = detail_programme.xpath("DateHeurelivraisonAuPlusTot")[0].text



                            try:
                                numero_identification = detail_programme.xpath("NumeroIdentificationAcheteur")[0].text
                            except IndexError:
                                numero_identification=""


                            try:
                                quantite=float(quantite)
                            except ValueError:
                                quantite=0
                            date_livraison=date_heure_livraison_au_plus_tot[0:8]
                            date_livraison=datetime.strptime(date_livraison, '%Y%m%d')
                            ligne = {
                                'quantite'             : quantite,
                                'type_commande'        : 'ferme',
                                'date_livraison'       : date_livraison,
                                'anomalie'             : '\n'.join(anomalie),
                                'point_dechargement'   : point_dechargement,
                                'numero_document'      : numero_document,  # N°UM de PSA 
                                'tg_number'            : tg_number,        # N°UM de Weidplas
                                'numero_identification': numero_identification,
                                'date_heure_livraison'             : date_heure_livraison,
                                'date_heure_livraison_au_plus_tot' : date_heure_livraison_au_plus_tot,
                                'code_routage'                     : code_routage,
                                'point_destination'                : point_destination,
                            }
                            res1.append(ligne)
                        val.update({'lignes':res1})
                        res.append(val)
                    else:
                        continue

        return res



    def get_data_GXS(self, attachment):
        res = []
        for obj in self:
            attachment=base64.decodebytes(attachment.datas)
            csvfile=attachment.split("\r")
            if len(csvfile)==1:
                csvfile=attachment.split("\n")
            tab=[]
            ct=0
            for row in csvfile:
                ct=ct+1
                lig=row.split(",")
                if len(lig)==29:
                    ref_article_client  = lig[14].strip()
                    num_commande_client = lig[16].strip()
                date_livraison = False
                type_commande  = "ferme"
                qt             = 0
                if len(lig)>6:
                    if lig[6].strip()=="PCE":
                        date_livraison=lig[0].strip()
                        d=datetime.strptime(date_livraison, '%Y%m%d')
                        date_livraison=d.strftime('%Y-%m-%d')
                        quantite=lig[4].strip()
                        try:
                            qt=float(quantite)
                        except ValueError:
                            continue
                        if lig[1].strip()=="Horizon Start Date":
                            type_commande  = "previsionnel"
                        val={
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        ligne = {
                            'quantite'      : qt,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes':[ligne]})
                        res.append(val)
        return res


    def get_data_John_Deere(self, attachment):
        res = []
        for obj in self:
            attachment=base64.decodebytes(attachment.datas)
            attachment=attachment.decode('iso-8859-1').encode('utf8').decode()
            csvfile=attachment.split("\n")
            tab=[]
            ct=0
            for row in csvfile:
                ct=ct+1
                lig=row.split(";")
                if len(lig)>=43:
                    type_commande=lig[0].strip()
                    if type_commande==u'Ferme' or type_commande==u'Prévisions':
                        ref_article_client  = lig[6].strip()
                        num_commande_client = lig[4].strip()
                        if type_commande==u'Ferme':
                            type_commande='ferme'
                        else:
                            type_commande='previsionnel'
                        date_livraison=lig[1].strip()
                        d=datetime.strptime(date_livraison, '%d/%m/%y')
                        date_livraison=d.strftime('%Y-%m-%d')
                        quantite=lig[7].strip()
                        try:
                            qt=float(quantite)
                        except ValueError:
                            qt=0
                        val={
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        ligne = {
                            'quantite'      : qt,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes':[ligne]})
                        res.append(val)
        return res






    def get_data_Motus(self, attachment):
        res = []
        lig=0
        for obj in self:
            filename = '/tmp/%s.xml' % attachment.id
            temp = open(filename, 'w+b')
            temp.write((base64.decodebytes(attachment.datas))) 
            temp.close()
            tree = ET.parse(filename)
            root = tree.getroot()
            for n1 in root:
                if n1.tag=='{urn:schemas-microsoft-com:office:spreadsheet}Worksheet':
                    for n2 in n1:
                        if n2.tag=='{urn:schemas-microsoft-com:office:spreadsheet}Table':
                            for n3 in n2:
                                lig=lig+1
                                if lig>1:
                                    col=0
                                    num_commande_client=''
                                    ref_article_client=''
                                    date_livraison=''
                                    type_commande=''
                                    quantite=''

                                    for n4 in n3:
                                        if n4.tag=='{urn:schemas-microsoft-com:office:spreadsheet}Cell':
                                            col=col+1
                                            for n5 in n4:
                                                if col==3:
                                                    num_commande_client=n5.text.strip()
                                                if col==5:
                                                    ref_article_client=n5.text.strip()
                                                if col==12:
                                                    type_commande=n5.text
                                                    if type_commande=='Firmed' or type_commande=='Partial':
                                                        type_commande='ferme'
                                                    else:
                                                        type_commande='previsionnel'
                                                if col==15:
                                                    quantite=n5.text
                                                    try:
                                                        qt=float(quantite)
                                                    except ValueError:
                                                        qt=0
                                                if col==17:
                                                    date_livraison=n5.text[:10]
                                                    date_livraison=datetime.strptime(date_livraison, '%Y-%m-%d')

                                    val={
                                        'num_commande_client' : num_commande_client,
                                        'ref_article_client'  : ref_article_client,
                                    }
                                    ligne = {
                                        'quantite'      : qt,
                                        'type_commande' : type_commande,
                                        'date_livraison': date_livraison,
                                    }
                                    val.update({'lignes':[ligne]})
                                    res.append(val)
        return res





    def get_data_Odoo(self, attachment):
        res = []
        for obj in self:
            csvfile=base64.b64decode(attachment.datas).decode("utf-8") 
            csvfile=csvfile.split("\n")
            tab=[]
            ct=0
            for row in csvfile:
                ct=ct+1
                lig=row.split("\t")
                if len(lig)==7:
                    # Recherche article
                    product=self.env['product.product'].search([
                        ('is_code'   , '=', lig[3]),
                    ])
                    anomalie=False
                    if len(product)==0:
                        anomalie=u'Article '+lig[3]+u' non trouvé'
                    # Recherche commande ouverte
                    order=self.env['sale.order'].search([
                        ('partner_id.is_code'    , '=', obj.partner_id.is_code),
                        ('is_article_commande_id', '=', product.id),
                        ('is_type_commande'      , '=', 'ouverte'),
                        ('state'                 , '=', 'draft'),
                    ])
                    if len(order)==0:
                        anomalie=u"Commande non trouvée pour l'article "+lig[3]
                    ref_article_client  = product.is_ref_client
                    num_commande_client = order.client_order_ref
                    val={
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                        'order_id'            : order.id,
                    }
                    type_commande="ferme"
                    if lig[5]=='prev':
                        type_commande="previsionnel"
                    date_livraison=lig[4].strip()
                    date_livraison=datetime.strptime(date_livraison, '%Y-%m-%d').date()
                    #date_livraison=d.strftime('%Y-%m-%d')
                    quantite=str(lig[6])
                    qt=0
                    try:
                        qt=float(quantite)
                    except ValueError:
                        qt=0
                    if qt!=0:
                        ligne = {
                            'quantite'      : qt,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                            'anomalie'      : anomalie
                        }
                        val.update({'lignes':[ligne]})
                        res.append(val)
        return res


    def get_data_plastika(self, attachment):
        res = []
        for obj in self:
            csvfile=base64.decodebytes(attachment.datas)
            csvfile=csvfile.split("\r\n")
            tab=[]
            ct=0
            for row in csvfile:
                ct=ct+1
                if ct>1:
                    lig=row.split("\t")
                    if len(lig)==3:
                        # Recherche article
                        product=self.env['product.product'].search([
                            ('is_code'   , '=', lig[0]),
                        ])
                        anomalie=False
                        if len(product)==0:
                            anomalie=u'Article '+lig[0]+u' non trouvé'
                        # Recherche commande ouverte
                        order=self.env['sale.order'].search([
                            ('partner_id.is_code'    , '=', obj.partner_id.is_code),
                            ('is_article_commande_id', '=', product.id),
                            ('is_type_commande'      , '=', 'ouverte'),
                            ('state'                 , '=', 'draft'),
                        ])
                        if len(order)==0:
                            anomalie=u"Commande non trouvée pour l'article "+lig[0]
                        ref_article_client  = product.is_ref_client
                        num_commande_client = order.client_order_ref
                        val={
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        type_commande="previsionnel"
                        date_livraison=lig[1].strip()
                        d=datetime.strptime(date_livraison, '%d/%m/%y')
                        date_livraison=d.strftime('%Y-%m-%d')
                        quantite=str(lig[2])
                        quantite=quantite.replace(",", ".")
                        quantite=quantite.replace(" ", "")
                        if quantite=='':
                            quantite=0
                        qt=0
                        try:
                            qt=float(quantite)
                        except ValueError:
                            qt=0
                        if qt!=0:
                            ligne = {
                                'quantite'      : qt,
                                'type_commande' : type_commande,
                                'date_livraison': date_livraison,
                                'anomalie'      : anomalie
                            }
                            val.update({'lignes':[ligne]})
                            res.append(val)
        return res


    # Nouvel EDI de THERMOR à partir du 15/05/2022
    def get_data_THERMOR(self, attachment):
        res = []
        for obj in self:
            #** Lecture du fichier xlsx ****************************************
            xlsxfile = base64.decodebytes(attachment.datas)
            path = '/tmp/THERMOR-'+str(obj.id)+'.xlsx'
            f = open(path,'wb')
            f.write(xlsxfile)
            f.close()
            type_fichier=False
            #*******************************************************************

            #** Ouverture du fichier *******************************************
            try:
                wb = openpyxl.load_workbook(filename = path)
                ws = wb.active
                cells = list(ws)
            except:
                raise ValidationError(u"Le fichier "+attachment.name+u" n'est pas un fichier xlsx valide")
            #*******************************************************************

            lig=0
            for row in ws.rows:
                if lig>0:
                    type_commande = 'previsionnel'
                    try:
                        quantite = float(cells[lig][7].value)
                    except ValueError:
                        quantite=0

                    try:
                        date_livraison = cells[lig][9].value.strftime('%Y-%m-%d')
                    except ValueError:
                        date_livraison = False

                    ref_article_client = cells[lig][5].value
                    order = self.env['sale.order'].search([
                        ('partner_id.is_code', '=', obj.partner_id.is_code),
                        ('is_ref_client'     , '=', ref_article_client)]
                    )
                    num_commande_client = "??"
                    if len(order):
                        num_commande_client = order[0].client_order_ref

                    val = {
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                    }
                    ligne = {
                        'quantite'      : quantite,
                        'type_commande' : type_commande,
                        'date_livraison': date_livraison,
                    }
                    val.update({'lignes': [ligne]})
                    res.append(val)
                lig+=1
        return res




    def get_data_Watts(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode('iso-8859-1')
            csvfile = csvfile.split("\r\n")
            csvfile = csv.reader(csvfile, delimiter='\t')
            for ct, lig in enumerate(csvfile):
                if len(lig)>=10 and ct>0:
                    ref_article_client = lig[1].strip()
                    order = self.env['sale.order'].search([
                        ('partner_id.is_code'   , '=', obj.partner_id.is_code),
                        ('is_ref_client', '=', ref_article_client),
                        ('is_type_commande'  , '=', 'ouverte'),
                    ])
                    num_commande_client = "??"
                    if len(order):
                        num_commande_client = order[0].client_order_ref
                    val = {
                        'num_commande_client' : num_commande_client,
                        'ref_article_client'  : ref_article_client,
                    }
                    try:
                        date_livraison=lig[8][0:10]
                        d=datetime.strptime(date_livraison, '%d/%m/%Y')
                        date_livraison=d.strftime('%Y-%m-%d')
                    except ValueError:
                        date_livraison=False
                    quantite = lig[9]
                    try:
                        quantite = float(quantite)
                    except ValueError:
                        quantite=0
                    type_commande="ferme"
                    if lig[0]==u'Prévision':
                        type_commande="previsionnel"
                    ligne = {
                        'quantite'      : quantite,
                        'type_commande' : type_commande,
                        'date_livraison': date_livraison,
                    }
                    val.update({'lignes': [ligne]})
                    res.append(val)
        return res





    def get_data_902580(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode()
            csvfile = csvfile.split("\n")
            csvfile = csv.reader(csvfile, delimiter='\t')
            for ct, lig in enumerate(csvfile):
                if len(lig)>=8:
                    quantite = lig[6]
                    try:
                        quantite = float(quantite)
                    except ValueError:
                        quantite=0
                    if quantite>0:
                        ref_article_client  = lig[0].strip()
                        num_commande_client = self.getNumCommandeClient(ref_article_client)
                        try:
                            date_livraison=lig[4]
                            d=datetime.strptime(date_livraison, '%d/%m/%Y')
                            date_livraison=d.strftime('%Y-%m-%d')
                        except ValueError:
                            date_livraison=False
                        type_commande="previsionnel"
                        if lig[5]=='F':
                            type_commande='ferme'
                        val = {
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        ligne = {
                            'quantite'      : quantite,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes': [ligne]})
                        res.append(val)
        return res


    def get_data_Valeo(self, attachment):
        res = []
        for obj in self:
            csvfile = base64.decodebytes(attachment.datas).decode()
            csvfile = csvfile.split("\n")
            csvfile = csv.reader(csvfile, delimiter=';')
            for ct, lig in enumerate(csvfile):
                if ct>0 and len(lig)>=42:
                    type_commande = 'previsionnel'
                    ref_article_client  = lig[17].strip() #[:-1]
                    num_commande_client = lig[30].strip()
                    date_livraison      = lig[40]
                    quantite            = lig[41]
                    try:
                        quantite = float(quantite)
                    except ValueError:
                        quantite=False
                    try:
                        date_livraison = datetime.strptime(date_livraison, '%d.%m.%Y')
                    except ValueError:
                        date_livraison = False
                    if quantite and date_livraison:
                        val = {
                            'num_commande_client' : num_commande_client,
                            'ref_article_client'  : ref_article_client,
                        }
                        ligne = {
                            'quantite'      : quantite,
                            'type_commande' : type_commande,
                            'date_livraison': date_livraison,
                        }
                        val.update({'lignes': [ligne]})
                        res.append(val)
        return res

